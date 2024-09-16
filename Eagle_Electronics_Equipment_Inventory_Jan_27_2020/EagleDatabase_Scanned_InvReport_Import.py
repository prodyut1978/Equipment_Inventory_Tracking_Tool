#Inventory Scanned Import
import os
from tkinter import*
import tkinter.messagebox
import EagleDatabase_BackEnd
import EagleDatabase_TransmittalOut_2
import tkinter.ttk as ttk
import tkinter as tk
import sqlite3
from tkinter.filedialog import asksaveasfile
from tkinter.filedialog import askopenfilename
from tkinter import simpledialog
import pandas as pd
import openpyxl
import csv
import time
import datetime

def ImportScannedFile():
    window = Tk()
    window.title("Import File Viewer")
    width = 1280
    height = 540
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x = (screen_width/2) - (width/2)
    y = (screen_height/2) - (height/2)
    window.geometry("%dx%d+%d+%d" % (width, height, x, y))
    window.resizable(0, 0)
    TableMargin = Frame(window)
    TableMargin.pack(side=TOP)
    TableMargin.pack(side=LEFT)
    scrollbarx = Scrollbar(TableMargin, orient=HORIZONTAL)
    scrollbary = Scrollbar(TableMargin, orient=VERTICAL)
    tree = ttk.Treeview(TableMargin, column=("column1", "column2", "column3", "column4", "column5", "column6", "column7", "column8", "column9", "column10", "column11"),
                                height=25, show='headings')
    scrollbary.config(command=tree.yview)
    scrollbary.pack(side=RIGHT, fill=Y)
    scrollbarx.config(command=tree.xview)
    scrollbarx.pack(side=BOTTOM, fill=X)

    tree.heading("#1", text="ManfSN", anchor=W)
    tree.heading("#2", text="Location", anchor=W)
    tree.heading("#3", text="Date", anchor=W)
    tree.heading("#4", text="AssetSN", anchor=W)
    tree.heading("#5", text="Category", anchor=W)
    tree.heading("#6", text="Manufacturer", anchor=W)
    tree.heading("#7", text="Model", anchor=W)
    tree.heading("#8", text="Description", anchor=W)            
    tree.heading("#9", text="Condition", anchor=W)
    tree.heading("#10", text="Origin", anchor=W)
    tree.heading("#11", text="Status", anchor=W)
    tree.column('#1', stretch=NO, minwidth=0, width=100)            
    tree.column('#2', stretch=NO, minwidth=0, width=120)
    tree.column('#3', stretch=NO, minwidth=0, width=90)
    tree.column('#4', stretch=NO, minwidth=0, width=80)
    tree.column('#5', stretch=NO, minwidth=0, width=90)
    tree.column('#6', stretch=NO, minwidth=0, width=90)
    tree.column('#7', stretch=NO, minwidth=0, width=90)
    tree.column('#8', stretch=NO, minwidth=0, width=105)
    tree.column('#9', stretch=NO, minwidth=0, width=80)
    tree.column('#10', stretch=NO, minwidth=0, width=80)
    tree.column('#11', stretch=NO, minwidth=0, width=80)
    tree.pack()

    def ImportScannedFile():
        name = askopenfilename(filetypes=[('Excel File', ('*.xlsx', '*.XLSX'))])
        if name:
            extensions = ['.xlsx', '.XLSX']
            if name.endswith(tuple(extensions)):
                wb   =  openpyxl.load_workbook(name)
                Sheet_Names = wb.sheetnames
                sheetN = 0
                dfList =[]
                for sheetN in range(len(Sheet_Names)):
                    sheet = wb[Sheet_Names[sheetN]]

                    Location = sheet.cell(row=3, column=3)
                    Location = (Location.value)
                    Date     = sheet.cell(row=5, column=3)
                    Date     = (Date.value)
                    
                    
                    df = pd.read_excel(name, sheet_name=sheetN, header = None, skiprows = 9)
                    df.rename(columns = {0:'Blank', 1:'Count', 2:'main_SN'},inplace = True)
                    data = pd.DataFrame(df)
                    data["location"]     = data.shape[0]*[Location]
                    data["datestamp"]    = data.shape[0]*[Date]
                    
                    data    = data.loc[:,['main_SN','location','datestamp']]
                    data = pd.DataFrame(data)
                    data = data.reset_index(drop=True)
                    dfList.append(data)

                concatDf = pd.concat(dfList, axis=0, ignore_index =True)
                data = pd.DataFrame(concatDf)
                if (data['main_SN'].duplicated().values.any() == True):
                    tkinter.messagebox.showinfo("Import File Error","Duplicate Manf_SN: Please Check the Imported File And Remove Duplicate Manf_SN")
                else:
                    if (data['main_SN'].isnull().values.any() == True)|(data['location'].isnull().values.any() == True)|(data['datestamp'].isnull().values.any() == True):
                        tkinter.messagebox.showinfo("Import File Error", "Manf_SN, Location And Date Can Not Be Empty ")
                    else:
                        tree.delete(*tree.get_children())
                        ImportTotalLBEntries.delete(0,END)
                        conn = sqlite3.connect("Eagle_Inventory.db")                       
                        data['datestamp'] = pd.to_datetime(data['datestamp']).dt.strftime("%Y-%m-%d")
                        data.to_sql('Eagle_Inventory_Scan_TEMP',conn, if_exists="replace", index=False)
                        conn.commit()
                        conn.close()                                        
                        for each_rec in range(len(data)):
                            tree.insert("", tk.END, values=list(data.loc[each_rec]))
                        ListBoxTotalImportEntries()
                
    def ListBoxTotalImportEntries():
        ImportTotalLBEntries.delete(0,END)
        Total_count = len(tree.get_children())
        ImportTotalLBEntries.insert(tk.END,Total_count)

    def PopulateScannedEntries():
        connMaster = sqlite3.connect("Eagle_Inventory.db")
        Complete_df_Master = pd.read_sql_query("SELECT * FROM Eagle_Inventory ORDER BY `main_SN` ASC ;", connMaster)
        data_Master = pd.DataFrame(Complete_df_Master)                   
        data_Master = data_Master.loc[:,['main_SN', 'asset_SN','catg','manuf','model', 'desc','condition','origin']]
        connMaster.commit()
        connMaster.close()

        connScannedImport = sqlite3.connect("Eagle_Inventory.db")
        Scanned_df_Import = pd.read_sql_query("SELECT * FROM Eagle_Inventory_Scan_TEMP ORDER BY `main_SN` ASC ;", connScannedImport)
        data_ScannedImport = pd.DataFrame(Scanned_df_Import)
        connScannedImport.commit()
        connScannedImport.close()
        LengthDF = len(data_ScannedImport)

        if LengthDF == 0:
            tkinter.messagebox.showinfo("Import file","Please Select the Import File to Merge with Master DB To Populate Empty Columns")
        else:
            tree.delete(*tree.get_children())
            ImportTotalLBEntries.delete(0,END)
            NonMatchedEntries.delete(0,END)

            Merge_Scanned_Master = pd.merge(data_ScannedImport, data_Master, on =['main_SN'] ,how ='left',
                               left_index = False, right_index = False, sort = True, indicator = True )
            

            def trans_MainSN_MISSING(x):
                if x   == 'both':
                    return 'Matched'
                elif x == 'right_only':
                    return 'Missing Equipment'
                elif x == 'left_only':
                    return 'Unknown ManufSN'
                else:
                    return x

            Merge_Scanned_Master['Status'] = Merge_Scanned_Master['_merge'].apply(trans_MainSN_MISSING)
            Merge_Scanned_Master = Merge_Scanned_Master.loc[:,
                            ['main_SN','location','datestamp','asset_SN','catg','manuf','model', 'desc','condition','origin','Status']]
            Merge_Scanned_Master["model"].fillna("Unknown", inplace = True)
            Merge_Scanned_Master["asset_SN"].fillna("Unknown", inplace = True)
            Merge_Scanned_Master["catg"].fillna("Unknown", inplace = True)
            Merge_Scanned_Master["manuf"].fillna("Unknown", inplace = True)
            Merge_Scanned_Master["desc"].fillna("Unknown", inplace = True)
            Merge_Scanned_Master["condition"].fillna("Unknown", inplace = True)
            Merge_Scanned_Master["origin"].fillna("Unknown", inplace = True)            
            Merge_Scanned_Master = Merge_Scanned_Master.reset_index(drop=True)
            Merge_Scanned_Master = pd.DataFrame(Merge_Scanned_Master)
            Merge_Scanned_Master_Count = len(Merge_Scanned_Master)
            connMerge_ToMasterDBTemp = sqlite3.connect("Eagle_Inventory.db")
            Merge_Scanned_Master.to_sql('Eagle_Inventory_Merged_temp',connMerge_ToMasterDBTemp, if_exists="replace", index=False)
            connMerge_ToMasterDBTemp.commit()
            connMerge_ToMasterDBTemp.close()

            Not_MatchedManfSN = Merge_Scanned_Master.loc[Merge_Scanned_Master.Status != 'Matched', 'main_SN': 'Status']
            Not_MatchedManfSN = Not_MatchedManfSN.reset_index(drop=True)
            Not_MatchedManfSN = pd.DataFrame(Not_MatchedManfSN)
            Not_MatchedManfSN_Count = len(Not_MatchedManfSN)
            
            for each_rec in range(len(Merge_Scanned_Master)):
                tree.insert("", tk.END, values=list(Merge_Scanned_Master.loc[each_rec]))
         
            ImportTotalLBEntries.insert(tk.END,Merge_Scanned_Master_Count)
            NonMatchedEntries.insert(tk.END,Not_MatchedManfSN_Count)

            if Not_MatchedManfSN_Count > 0:
                tkinter.messagebox.showinfo("Message For Import","There are unknown Equipment Manufacturer Serial Number in the Import File")
            else:
                tkinter.messagebox.showinfo("Message For Import","There are no unknown Manufacturer Serial Number Equipment in the Import File")
                
    def PopulateNonMatchedScannedEntries():
        NonMatchedEntries.delete(0,END)
        con= sqlite3.connect("Eagle_Inventory.db")
        Merge_Scanned_Master_df = pd.read_sql_query("SELECT * FROM Eagle_Inventory_Merged_temp ORDER BY `catg` ASC ;", con)
        con.commit()
        con.close()
        
        Not_MatchedManfSN = Merge_Scanned_Master_df.loc[Merge_Scanned_Master_df.Status != 'Matched', 'main_SN': 'Status']
        Not_MatchedManfSN = Not_MatchedManfSN.reset_index(drop=True)
        Not_MatchedManfSN = pd.DataFrame(Not_MatchedManfSN)
        Not_MatchedManfSN_Count = len(Not_MatchedManfSN)
        NonMatchedEntries.insert(tk.END,Not_MatchedManfSN_Count)
            
        def InventoryRec(event):
            for nm in tree1.selection():
                sd = tree1.item(nm, 'values')

                lblMain_SNEntries.delete(0,END)
                lblMain_SNEntries.insert(tk.END,sd[0])

                lblLocationEntries.delete(0,END)
                lblLocationEntries.insert(tk.END,sd[1])

                lblDateEntries.delete(0,END)
                lblDateEntries.insert(tk.END,sd[2])

                lblAssetSNEntries.delete(0,END)
                lblAssetSNEntries.insert(tk.END,sd[3])
                
                lblCatgEntries.delete(0,END)
                lblCatgEntries.insert(tk.END,sd[4])
                
                lblManufEntries.delete(0,END)
                lblManufEntries.insert(tk.END,sd[5])
                
                lblModelEntries.delete(0,END)
                lblModelEntries.insert(tk.END,sd[6])
                
                lblDescriptionEntries.delete(0,END)
                lblDescriptionEntries.insert(tk.END,sd[7])                
                
                lblConditionEntries.delete(0,END)
                lblConditionEntries.insert(tk.END,sd[8])
                
                lblOriginEntries.delete(0,END)
                lblOriginEntries.insert(tk.END,sd[9])
                    
        window = Tk()
        window.title("Input Entry For Non Matched Manuf SN Equipment")
        window.config(bg="ghost white")
        width = 1050
        height = 520
        screen_width = window.winfo_screenwidth()
        screen_height = window.winfo_screenheight()
        x = (screen_width/2) - (width/2)
        y = (screen_height/2) - (height/2)
        window.geometry("%dx%d+%d+%d" % (width, height, x, y))
        window.resizable(0, 0)
        window.grid()
        TitFrame = Frame(window, bd = 2, padx= 5, pady= 4, relief = RIDGE)
        TitFrame.pack(side = TOP)
        InputHeader = Label(TitFrame, font=('aerial', 12, 'bold'), text="Input Entry For Equipment")
        InputHeader.grid()

        TableMargin = LabelFrame(window, bd = 1, width = 490, height = 400, padx= 6, pady= 10,relief = RIDGE,
                                           bg = "Ghost White",font=('aerial', 15, 'bold'))
        TableMargin.pack(side=TOP)
        TableMargin.pack(side=RIGHT)
        scrollbarx = Scrollbar(TableMargin, orient=HORIZONTAL)
        scrollbary = Scrollbar(TableMargin, orient=VERTICAL)
        tree1 = ttk.Treeview(TableMargin, column=("column1", "column2", "column3", "column4", "column5", "column6", "column7", "column8", "column9", "column10", "column11"),
                                    height=18, show='headings')
        scrollbary.config(command=tree.yview)
        scrollbary.pack(side=RIGHT, fill=Y)
        scrollbarx.config(command=tree.xview)
        scrollbarx.pack(side=BOTTOM, fill=X)

        tree1.heading("#1", text="ManSN", anchor=W)
        tree1.heading("#2", text="Loc", anchor=W)
        tree1.heading("#3", text="Date", anchor=W)
        tree1.heading("#4", text="AstSN", anchor=W)
        tree1.heading("#5", text="Catagory", anchor=W)
        tree1.heading("#6", text="Manuf", anchor=W)
        tree1.heading("#7", text="Model", anchor=W)
        tree1.heading("#8", text="Desc", anchor=W)            
        tree1.heading("#9", text="Condition", anchor=W)
        tree1.heading("#10", text="Origin", anchor=W)
        tree1.heading("#11", text="Status", anchor=W)
        tree1.column('#1', stretch=NO, minwidth=0, width=50)            
        tree1.column('#2', stretch=NO, minwidth=0, width=50)
        tree1.column('#3', stretch=NO, minwidth=0, width=50)
        tree1.column('#4', stretch=NO, minwidth=0, width=50)
        tree1.column('#5', stretch=NO, minwidth=0, width=50)
        tree1.column('#6', stretch=NO, minwidth=0, width=50)
        tree1.column('#7', stretch=NO, minwidth=0, width=50)
        tree1.column('#8', stretch=NO, minwidth=0, width=50)
        tree1.column('#9', stretch=NO, minwidth=0, width=50)
        tree1.column('#10', stretch=NO, minwidth=0, width=50)
        tree1.column('#11', stretch=NO, minwidth=0, width=50)
        tree1.pack()
        tree1.bind('<<TreeviewSelect>>',InventoryRec)
        for each_rec in range(len(Not_MatchedManfSN)):
            tree1.insert("", tk.END, values=list(Not_MatchedManfSN.loc[each_rec]))

        DataFrameLEFT = LabelFrame(window, bd = 1, width = 490, height = 400, padx= 6, pady= 10,relief = RIDGE,
                                           bg = "Ghost White",font=('aerial', 15, 'bold'))
        DataFrameLEFT.place(x=0,y=59)

        lblCatg = Label(DataFrameLEFT, font=('aerial', 10, 'bold'), text = "1. Category:", padx =10, pady= 10, bg = "Ghost White")
        lblCatg.grid(row =0, column = 0, sticky =W)
        lblCatgEntries  = Entry(DataFrameLEFT, font=('aerial', 12, 'bold'),textvariable= StringVar(), width = 30, bd=2)
        lblCatgEntries.grid(row =0, column = 1)

        lblManuf = Label(DataFrameLEFT, font=('aerial', 10, 'bold'), text = "2. Manufacturer:", padx =10, pady= 10, bg = "Ghost White")
        lblManuf.grid(row =1, column = 0, sticky =W)
        lblManufEntries  = Entry(DataFrameLEFT, font=('aerial', 12, 'bold'),textvariable= StringVar(), width = 30, bd=2)
        lblManufEntries.grid(row =1, column = 1)

        lblModel = Label(DataFrameLEFT, font=('aerial', 10, 'bold'), text = "3. Model Name:", padx =10, pady= 10, bg = "Ghost White")
        lblModel.grid(row =2, column = 0, sticky =W)
        lblModelEntries  = Entry(DataFrameLEFT, font=('aerial', 12, 'bold'),textvariable= StringVar(), width = 30, bd=2)
        lblModelEntries.grid(row =2, column = 1)

        lblMain_SN = Label(DataFrameLEFT, font=('aerial', 10, 'bold'), text = "4. Main SN:", padx =10, pady= 10, bg = "Ghost White")
        lblMain_SN.grid(row =3, column = 0, sticky =W)
        lblMain_SNEntries  = Entry(DataFrameLEFT, font=('aerial', 12, 'bold'),textvariable= StringVar(), width = 30, bd=2)
        lblMain_SNEntries.grid(row =3, column = 1)

        lblDescription = Label(DataFrameLEFT, font=('aerial', 10, 'bold'), text = "5. Description:", padx =10, pady= 10, bg = "Ghost White")
        lblDescription.grid(row =4, column = 0, sticky =W)
        lblDescriptionEntries  = Entry(DataFrameLEFT, font=('aerial', 12, 'bold'),textvariable= StringVar(), width = 30, bd=2)
        lblDescriptionEntries.grid(row =4, column = 1)

        lblAssetSN = Label(DataFrameLEFT, font=('aerial', 10, 'bold'), text = "6. Asset SN:", padx =10, pady= 10, bg = "Ghost White")
        lblAssetSN.grid(row =5, column = 0, sticky =W)
        lblAssetSNEntries  = Entry(DataFrameLEFT, font=('aerial', 12, 'bold'),textvariable= StringVar(), width = 30, bd=2)
        lblAssetSNEntries.grid(row =5, column = 1)

        lblDate = Label(DataFrameLEFT, font=('aerial', 10, 'bold'), text = "7. Date (yyyy-mm-dd):", padx =10, pady= 10, bg = "Ghost White")
        lblDate.grid(row =6, column = 0, sticky =W)
        lblDateEntries  = Entry(DataFrameLEFT, font=('aerial', 12, 'bold'),textvariable= StringVar(), width = 30, bd=2)
        lblDateEntries.grid(row =6, column = 1)

        lblLocation = Label(DataFrameLEFT, font=('aerial', 10, 'bold'), text = "8. Location:", padx =10, pady= 10, bg = "Ghost White")
        lblLocation.grid(row =7, column = 0, sticky =W)
        lblLocationEntries  = Entry(DataFrameLEFT, font=('aerial', 12, 'bold'),textvariable= StringVar(), width = 30, bd=2)
        lblLocationEntries.grid(row =7, column = 1)

        lblCondition = Label(DataFrameLEFT, font=('aerial', 10, 'bold'), text = "9. Condition:", padx =10, pady= 10, bg = "Ghost White")
        lblCondition.grid(row =8, column = 0, sticky =W)
        lblConditionEntries  = Entry(DataFrameLEFT, font=('aerial', 12, 'bold'),textvariable= StringVar(), width = 30, bd=2)
        lblConditionEntries.grid(row =8, column = 1)

        lblOrigin = Label(DataFrameLEFT, font=('aerial', 10, 'bold'), text = "10. Origin:", padx =10, pady= 10, bg = "Ghost White")
        lblOrigin.grid(row =9, column = 0, sticky =W)
        lblOriginEntries  = Entry(DataFrameLEFT, font=('aerial', 12, 'bold'),textvariable= StringVar(), width = 30, bd=2)
        lblOriginEntries.grid(row =9, column = 1)

        def ClearAll():
            lblCatgEntries.delete(0,END)
            lblManufEntries.delete(0,END)
            lblModelEntries.delete(0,END)                                                
            lblMain_SNEntries.delete(0,END)
            lblDescriptionEntries.delete(0,END)
            lblAssetSNEntries.delete(0,END)
            lblDateEntries.delete(0,END)
            lblLocationEntries.delete(0,END)
            lblConditionEntries.delete(0,END)
            lblOriginEntries.delete(0,END)

        def iExit():
            window.destroy()

        def AddData():
            if(len(lblMain_SNEntries.get())!=0) & (len(lblAssetSNEntries.get())!=0):
                EagleDatabase_BackEnd.addInvRec(lblCatgEntries.get(), lblManufEntries.get(), lblModelEntries.get(),
                                                lblMain_SNEntries.get(),lblDescriptionEntries.get(), lblAssetSNEntries.get(), lblDateEntries.get(),
                                                lblLocationEntries.get(), lblConditionEntries.get(), lblOriginEntries.get())
                tree1.delete(*tree1.get_children())
                tree1.insert("", tk.END,values=(lblCatgEntries.get(), lblManufEntries.get(), lblModelEntries.get(),
                                                lblMain_SNEntries.get(),lblDescriptionEntries.get(), lblAssetSNEntries.get(),
                                                lblDateEntries.get(), lblLocationEntries.get(),
                                                lblConditionEntries.get(), lblOriginEntries.get()))
                PopulateScannedEntries()
            else:
                tkinter.messagebox.showinfo("Add Error","Manufacture and Asset SN can not be empty")

        btnExit = Button(window, text="Exit", font=('aerial', 10, 'bold'), height =1, width=8, bd=2, command = iExit)
        btnExit.place(x=2,y=490)
        btnClear = Button(window, text="Clear Entries", font=('aerial', 10, 'bold'), height =1, width=10, bd=2, command = ClearAll)
        btnClear.place(x=80,y=490)
        btnAdd = Button(window, text="Add Entry", font=('aerial', 10, 'bold'), height =1, width=8, bd=2, command = AddData)
        btnAdd.place(x=372,y=490)

    def DeleteMergedSelected():
        iDelete = tkinter.messagebox.askyesno("Delete Entry From Transmittal", "Confirm if you want to Delete")
        if iDelete >0:
            conn = sqlite3.connect("Eagle_Inventory.db")
            cur = conn.cursor()
            ImportTotalLBEntries.delete(0,END)
            NonMatchedEntries.delete(0,END)
            for selected_item in tree.selection():
                cur.execute("DELETE FROM Eagle_Inventory_Merged_temp WHERE main_SN =? AND model = ? ", (tree.set(selected_item, '#1'), tree.set(selected_item, '#7'),))
                conn.commit()
                tree.delete(selected_item)
            Total_count = len(tree.get_children())
            ImportTotalLBEntries.insert(tk.END,Total_count)
            conn.commit()
            conn.close()
            NonMatchedReCalculation()
        return
                    
    def NonMatchedReCalculation():
        NonMatchedEntries.delete(0,END)
        con= sqlite3.connect("Eagle_Inventory.db")
        Merge_Scanned_Master_df = pd.read_sql_query("SELECT * FROM Eagle_Inventory_Merged_temp ORDER BY `catg` ASC ;", con)
        con.commit()
        con.close()
        
        Not_MatchedManfSN = Merge_Scanned_Master_df.loc[Merge_Scanned_Master_df.Status != 'Matched', 'main_SN': 'Status']
        Not_MatchedManfSN = Not_MatchedManfSN.reset_index(drop=True)
        Not_MatchedManfSN = pd.DataFrame(Not_MatchedManfSN)
        Not_MatchedManfSN_Count = len(Not_MatchedManfSN)
        NonMatchedEntries.insert(tk.END,Not_MatchedManfSN_Count)


    def UpdateToMasterDB():
        con= sqlite3.connect("Eagle_Inventory.db")
        Imported_df = pd.read_sql_query("SELECT * FROM Eagle_Inventory_Merged_temp ORDER BY `catg` ASC ;", con)
        Imported_df = Imported_df.loc[:,
                        ['catg', 'manuf', 'model', 'main_SN', 'desc', 'asset_SN', 'datestamp', 'location', 'condition', 'origin']]
        Imported_df = Imported_df.reset_index(drop=True)
        Imported_df = pd.DataFrame(Imported_df)
        LengthDF = len(Imported_df)
        con.commit()
        con.close()
        if LengthDF == 0:
            tkinter.messagebox.showinfo("Import file","Please Select Import Scanned File > Merge All Columns To Update Import In Master DB")
        else:
            iSubmit = tkinter.messagebox.askyesno("Entries Submit to Master DB", "Confirm if you want to Submit the Imported Entries to Master DB")
            if iSubmit >0:
                if (Imported_df['main_SN'].isnull().values.any() == True)|(Imported_df['asset_SN'].isnull().values.any() == True)|(Imported_df['catg'].isnull().values.any() == True)|(Imported_df['manuf'].isnull().values.any() == True)|(Imported_df['model'].isnull().values.any() == True):
                    tkinter.messagebox.showinfo("Import File Message", "MainSN,Asset SN, Category, Manufacturer Model Cannot be Empty ")
                else:
                    con= sqlite3.connect("Eagle_Inventory.db")
                    cur=con.cursor()
                    Imported_df.to_sql('Eagle_Inventory_transmittalOut2',con, if_exists="replace", index=False)
                    cur.execute("DELETE FROM Eagle_Inventory WHERE EXISTS (SELECT * FROM Eagle_Inventory_Merged_temp WHERE Eagle_Inventory.main_SN = Eagle_Inventory_Merged_temp.main_SN and Eagle_Inventory.model = Eagle_Inventory_Merged_temp.model)")                        
                    cur.execute("INSERT INTO Eagle_Inventory (catg, manuf, model, main_SN, desc, asset_SN,\
                                    datestamp, location, condition, origin) SELECT catg, manuf, model, main_SN, desc, asset_SN, datestamp, location, condition, origin FROM Eagle_Inventory_Merged_temp")
                    cur.execute("DELETE FROM Eagle_Inventory_Merged_temp")
                    cur.execute("DELETE FROM Eagle_Inventory_Scan_TEMP")
                    time.sleep(2)
                    con.commit()
                    con.close()          
                    tree.delete(*tree.get_children())
                    tkinter.messagebox.showinfo("Submitted to Inventory Database(DB)","You have Submitted a Record to Master Inventory Database(DB)")                
            return
            
    def Exit():
        window.destroy()

    def ClearAll():
            tree.delete(*tree.get_children())
            ImportTotalLBEntries.delete(0,END)
            con= sqlite3.connect("Eagle_Inventory.db")
            cur=con.cursor()
            try:
                cur.execute("DELETE FROM Eagle_Inventory_Scan_TEMP")
            except:
                tkinter.messagebox.showinfo("Clear DB","Import Database Are Already Cleared")

            try:
                cur.execute("DELETE FROM Eagle_Inventory_Merged_temp")
            except:
                tkinter.messagebox.showinfo("Clear DB","Merged Database Are Already Cleared")

            try:
                cur.execute("DELETE FROM Eagle_Inventory_transmittalOut2")
            except:
                tkinter.messagebox.showinfo("Clear DB","Transmittal Out Database Are Already Cleared")

            con.commit()
            con.close()
            tkinter.messagebox.showinfo("Clear DB","All Temp Database are Cleared")  


    def GenerateTransmittalOut2():
        iSubmit = tkinter.messagebox.askyesno("Generate Transmittal Out",
                "Confirm if you want to Generate Transmittal Out")
        if iSubmit >0:        
            EagleDatabase_TransmittalOut_2.GenerateTransmittalOutSecondOption()
            Exit()
        return


    def DisplayMergedTemp():
        con= sqlite3.connect("Eagle_Inventory.db")
        Merge_Scanned_df = pd.read_sql_query("SELECT * FROM Eagle_Inventory_Merged_temp ORDER BY `catg` ASC ;", con)
        con.commit()
        con.close()
        Merge_Scanned_df = Merge_Scanned_df.reset_index(drop=True)
        for each_rec in range(len(Merge_Scanned_df)):
            tree.insert("", tk.END, values=list(Merge_Scanned_df.loc[each_rec]))

    def UpdateSlectedCondition():
        dfList =[] 
        for child in tree.get_children():
            df = tree.item(child)["values"]
            dfList.append(df)
        ListBox_DF = pd.DataFrame(dfList)
        if len(ListBox_DF)>0:
            SelectionTree = tree.selection()
            if len(SelectionTree)>0:
                iUpdateSlectedCond = tkinter.messagebox.askyesno("Update Condition in Database", "Confirm if you want to Update Condition")
                if iUpdateSlectedCond >0:
                    conn = sqlite3.connect("Eagle_Inventory.db")
                    cur = conn.cursor()
                    application_window = Tk()
                    
                    width = 5
                    height = 5
                    screen_width = application_window.winfo_screenwidth()
                    screen_height = application_window.winfo_screenheight()
                    x = (screen_width/2) - (width/2)
                    y = (screen_height/2) - (height/2)
                    application_window.geometry("%dx%d+%d+%d" % (width, height, x, y))
                    application_window.resizable(0, 0)
                    application_window.lift()

                    cond_update = simpledialog.askstring("Input Updated Condition", "What is your updated condition?",
                                                                             parent=application_window)
                    if cond_update is None:
                        tkinter.messagebox.showinfo("Update Error","Please Input Updated Condition")
                        application_window.destroy()
                    else:
                        for selected_item in tree.selection():
                            cur.execute("UPDATE Eagle_Inventory_Merged_temp SET condition =? WHERE main_SN =? AND model = ? ", (cond_update, tree.set(selected_item, '#1'),tree.set(selected_item, '#7')))
                            conn.commit()                        
                        conn.commit()
                        conn.close()
                        application_window.destroy()
                        tree.delete(*tree.get_children())
                        DisplayMergedTemp()
                        tkinter.messagebox.showinfo("Update Condition Message","Updated Condition Complete")
                    application_window.mainloop()
                return
            else:
                tkinter.messagebox.showinfo("Update Error","Please Select Entries To Update Condition")            
        else:
            tkinter.messagebox.showinfo("Update Error","Please Select Entries To Update Condition")
                    

    btnExit = Button(window, text="Exit", font=('aerial', 9, 'bold'), height =1, width=10, bd=4, command = Exit)
    btnExit.place(x=1194,y=508)

    btnClearAll = Button(window, text="ClearAll", font=('aerial', 9, 'bold'), height =1, width=10, bd=4, command = ClearAll)
    btnClearAll.place(x=1194,y=475)

    L1Import = Label(window, text = "1: Import External Scanned File", font=("arial", 10,'bold'),bg = "ghost white").place(x=1040,y=20)
    btnImportScan = Button(window, text="Import Scanned File", font=('aerial', 9, 'bold'), height =1, width=16, bd=4, command = ImportScannedFile)
    btnImportScan.place(x=1040,y=45)
    ImportTotalLBEntries  = Entry(window, font=('aerial', 12, 'bold'),textvariable= IntVar(), width = 7)
    ImportTotalLBEntries.place(x=1180,y=48)


    L2Populate = Label(window, text = "2: Merge All Empty Columns", font=("arial", 10,'bold'),bg = "ghost white").place(x=1040,y=110)
    btnPopulate = Button(window, text="Merge All Columns", font=('aerial', 9, 'bold'), height =1, width=16, bd=4, command = PopulateScannedEntries)
    btnPopulate.place(x=1040,y=135)
    btnDeleteMergeSelected = Button(window, text="Delete Selected", font=('aerial', 9, 'bold'), height =1, width=13, bd=2, command = DeleteMergedSelected)
    btnDeleteMergeSelected.place(x=1175,y=138)

    L3NotMatchedSN = Label(window, text = "3: View Not Matched ManfSN", font=("arial", 10,'bold'),bg = "ghost white").place(x=1040,y=210)
    btnViewNonMatched = Button(window, text="View Non Matched", font=('aerial', 9, 'bold'), height =1, width=16, bd=4, command = PopulateNonMatchedScannedEntries)
    btnViewNonMatched.place(x=1040,y=235)
    NonMatchedEntries  = Entry(window, font=('aerial', 12, 'bold'),textvariable= IntVar(), width = 7)
    NonMatchedEntries.place(x=1180,y=238)

    L3UpdateDB = Label(window, text = "4. Update Import To Master DB", font=("arial", 10,'bold'),bg = "ghost white").place(x=1040,y=305)
    btnUpdateCondition = Button(window, text="Update Condition", font=('aerial', 9, 'bold'), height =1, width=16, bd=4, command = UpdateSlectedCondition)
    btnUpdateCondition.place(x=1040,y=330)
    L3UpdateCond = Label(window, text = "If Required", font=("arial", 10,'bold'),bg = "ghost white").place(x=1180,y=330)
    btnUpdateImportToDB = Button(window, text="Update Master DB", font=('aerial', 9, 'bold'), height =1, width=16, bd=4, command = UpdateToMasterDB)
    btnUpdateImportToDB.place(x=1040,y=365)

    L3MakeTransmittalOut = Label(window, text = "5. Generate Transmittal Out", font=("arial", 10,'bold'), bg = "ghost white").place(x=1040,y=430)
    btnGenerateTransmittalOut = Button(window, text="Transmittal Out", font=('aerial', 9, 'bold'), height =1, width=16, bd=4, command = GenerateTransmittalOut2)
    btnGenerateTransmittalOut.place(x=1040,y=455)











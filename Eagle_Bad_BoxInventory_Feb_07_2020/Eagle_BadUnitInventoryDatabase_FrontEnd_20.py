#Front End
import os
from tkinter import*
import tkinter.messagebox
import Eagle_BadGSRInventoryDatabase_BackEnd
import GenerateReceivingAddress
import tkinter.ttk as ttk
import tkinter as tk
import sqlite3
from tkinter.filedialog import asksaveasfile
from tkinter.filedialog import askopenfilename
from tkinter import simpledialog
from tkinter import filedialog
import pandas as pd
import openpyxl
import csv
import time
import datetime
import win32com.client

Default_Date_today   = datetime.date.today()

class Inventory:
    
    def __init__(self,root):
        self.root =root
        self.root.title ("Eagle Equipment Inventory")
        self.root.geometry("1355x690+0+0")
        self.root.config(bg="cadet blue")
        self.root.resizable(0, 0)
        

##  ----------------- Define Variables-------------
        
        BatchNumber      = StringVar()
        JobName          = StringVar()
        CrewNumber       = StringVar()
        Location         = StringVar()
        Date             = StringVar(self.root, value=Default_Date_today)
        Unit_SN          = StringVar()
        DeviceType       = StringVar(self.root, value='GSRx3')                
        Opened           = StringVar(self.root, value='No')
        FaultFound       = StringVar()
        SEARCH           = StringVar()
        TOTALE           = IntVar()
        SEARCHM          = StringVar()
        TOTALLB          = IntVar()
        TOTALD           = IntVar()

        ##  ----------------- Define Functions------------
        
        def ExportCompleteDB():
            conn = sqlite3.connect("Eagle_BadGSRInventoryDatabase.db")
            Complete_df = pd.read_sql_query("select * from Eagle_BadGSRInventoryDatabase ;", conn)
            Export_Database = pd.DataFrame(Complete_df)
            filename = tkinter.filedialog.asksaveasfilename(initialdir = "/",title = "Select file" ,\
                       defaultextension='.xlsx', filetypes = (("Excel file",".xlsx"),("TXT file",".txt")))
            if filename:
                if filename.endswith('.txt'):
                    with open(filename, 'w') as file:
                        Export_Database.to_csv(file,index=None)
                    file.close
                    tkinter.messagebox.showinfo("DB Export","DB Saved as TXT")
                else:
                    with pd.ExcelWriter(filename) as file:
                        Export_Database.to_excel(file,sheet_name='InventoryDB',index=False)
                    file.close
                    tkinter.messagebox.showinfo("DB Export","DB Saved as Excel")                    
                        
            conn.commit()
            conn.close()
        
        def update():
            if(len(BatchNumber.get())!=0) & (len(Unit_SN.get())!=0) & (len(JobName.get())!=0) & (len(Date.get())!=0): 
                conn = sqlite3.connect("Eagle_BadGSRInventoryDatabase.db")
                cur = conn.cursor()
                for selected_item in tree.selection():
                    cur.execute("DELETE FROM Eagle_BadGSRInventoryDatabase WHERE BatchNumber =? AND JobName=? AND \
                                CrewNumber =? AND Location =? AND Date =? AND Unit_SN =? AND DeviceType =? AND Opened =? AND FaultFound =? ",\
                                (tree.set(selected_item, '#1'), tree.set(selected_item, '#2'),tree.set(selected_item, '#3'),\
                                 tree.set(selected_item, '#4'),tree.set(selected_item, '#5'),tree.set(selected_item, '#6'),\
                                 tree.set(selected_item, '#7'),tree.set(selected_item, '#8'),tree.set(selected_item, '#9'),))
                    conn.commit()
                    tree.delete(selected_item)
                    conn.close()

            if(len(BatchNumber.get())!=0) & (len(Unit_SN.get())!=0) & (len(JobName.get())!=0) & (len(Date.get())!=0): 
                Eagle_BadGSRInventoryDatabase_BackEnd.addInvRec(BatchNumber.get(), JobName.get(), CrewNumber.get(), Location.get(),Date.get(), Unit_SN.get(), DeviceType.get(), Opened.get(), FaultFound.get())
                tree.delete(*tree.get_children())
                tree.insert("", tk.END,values=(BatchNumber.get(), JobName.get(), CrewNumber.get(), Location.get(),Date.get(), Unit_SN.get(), DeviceType.get(), Opened.get(), FaultFound.get()))
            else:
                tkinter.messagebox.showinfo("Update Error","Batch Number, Job Name, Date and GSR Serial Number entry can not be empty")
        

        def ClearAllView():
               tree.delete(*tree.get_children())
               Reset_TotalDBCount()
               Reset_TotalDuplicateCount()
        
        def KeySearch():
            if SEARCH.get() != "":
                tree.delete(*tree.get_children())
                conn = sqlite3.connect("Eagle_BadGSRInventoryDatabase.db")
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM `Eagle_BadGSRInventoryDatabase` WHERE `BatchNumber`= ? COLLATE NOCASE OR JobName = ? COLLATE NOCASE OR Unit_SN = ? COLLATE NOCASE OR FaultFound = ? COLLATE NOCASE OR Location = ? COLLATE NOCASE",\
                               (self.txtKeySearch.get(), self.txtKeySearch.get(), self.txtKeySearch.get(), self.txtKeySearch.get(),self.txtKeySearch.get(),))                
                fetch = cursor.fetchall()
                for data in fetch:
                    tree.insert('', 'end', values=(data))
                cursor.close()
                conn.close()
            

        
        def ExportSelectedListBoxView():
            filename = tkinter.filedialog.asksaveasfilename(initialdir = "/",title = "Select file" ,\
                       defaultextension='.csv', filetypes = (("CSV file",".csv"),("Text file",".txt")))
            if filename:
                    if filename.endswith('.csv'):
                        with open(filename, 'w') as file:
                            file.write('BatchNumber' + ',' + 'JobName' + ',' + 'CrewNumber' + ',' + 'Location' + ',' +  'Date ' + ',' + 'Unit_SN' +\
                                       ',' + 'DeviceType' + ',' + 'Opened' + ',' + 'FaultFound' +'\n')
                            for item in tree.selection():
                                list_item = (tree.item(item, 'values'))
                                x1= list_item[0]
                                x2= list_item[1]
                                x3= list_item[2]
                                x4= list_item[3]
                                x5= list_item[4]
                                x6= list_item[5]
                                x7= list_item[6]
                                x8= list_item[7]
                                x9= list_item[8]                                
                                file.write( x1 + ',' + x2 + ',' + x3 + ',' + x4 + ',' + x5 + ',' + x6 + ',' + x7 + ',' + x8 +  ',' + x9 + '\n')
                        file.close
                        tkinter.messagebox.showinfo("Save file","File Saved as CSV")

                    else:
                        with open(filename, 'w') as file:
                            file.write('BatchNumber' + ',' + 'JobName' + ',' + 'CrewNumber' + ',' + 'Location' + ',' +  'Date ' + ',' + 'Unit_SN' +\
                                       ',' + 'DeviceType' + ',' + 'Opened' + ',' + 'FaultFound' +'\n')
                            for item in tree.selection():
                                list_item = (tree.item(item, 'values'))
                                x1= list_item[0]
                                x2= list_item[1]
                                x3= list_item[2]
                                x4= list_item[3]
                                x5= list_item[4]
                                x6= list_item[5]
                                x7= list_item[6]
                                x8= list_item[7]
                                x9= list_item[8]
                                file.write( x1 + ',' + x2 + ',' + x3 + ',' + x4 + ',' + x5 + ',' + x6 + ',' + x7 + ',' + x8 +  ',' + x9 + '\n')                                
                        file.close
                        tkinter.messagebox.showinfo("Save file","File Saved as TEXT")


        def ExportListboxEntries():
            dfList =[] 
            for child in tree.get_children():
                df = tree.item(child)["values"]
                dfList.append(df)
            Export_DF = pd.DataFrame(dfList)
            Export_DF.rename(columns = {0:'BatchNumber', 1:'JobName', 2:'CrewNumber', 3:'Location', 4:'Date',
                                          5: 'Unit_SN', 6:'DeviceType', 7:'Opened', 8:'FaultFound'},inplace = True)
            data_SortByUnit_SN = Export_DF.sort_values(by =['Unit_SN'])
            
            filename = tkinter.filedialog.asksaveasfilename(initialdir = "/",title = "Export file" ,
                       defaultextension='.xlsx', filetypes = (("Excel file",".xlsx"),("Excel file",".xlsx")))
            if filename:
                if filename.endswith('.xlsx'):
                    with pd.ExcelWriter(filename) as file:
                        data_SortByUnit_SN.to_excel(file,sheet_name='BadBox_Inventory',index=False)
                    file.close
                    tkinter.messagebox.showinfo("ListBox Entries Export"," ListBox Entries Exported as Excel")
            tree.delete(*tree.get_children())

        def Transmittal_ConvertToPDF():
            o = win32com.client.Dispatch("Excel.Application")
            o.Visible = False
            wb_path = tkinter.filedialog.askopenfilename(initialdir = "/" ,title = "Select Exported Transmittal Out Excel File" ,
                                                   defaultextension='.xlsx', filetypes = (("Excel file",".xlsx"),("Excel file",".xlsx")))
            wb = o.Workbooks.Open(wb_path)
            ws_index_list = [1] #say you want to print these sheets
            FILEOPENOPTIONS = dict(defaultextension=".pdf", initialdir="/", title = "Name the PDF File to Save" ,
                                       filetypes=[('pdf file', '*.pdf')])
            path_to_pdf  = filedialog.asksaveasfilename(**FILEOPENOPTIONS)
            wb.WorkSheets(ws_index_list).Select()
            wb.ActiveSheet.ExportAsFixedFormat(0, path_to_pdf)
            wb.close
            tkinter.messagebox.showinfo("Transmittal To PDF"," Transmittal Out Export Report Saved as PDF")


        def UpdateMasterDB():
            iUpdate = tkinter.messagebox.askyesno("Update Master DB",
                    "Confirm if you want to Update Master DB Removing Duplicated CaseSrNo With New One")
            if iUpdate >0:
                tree.delete(*tree.get_children())
                self.txtDuplicatedEntries.delete(0,END)
                self.txtTotalEntries.delete(0,END)
                conn = sqlite3.connect("Eagle_BadGSRInventoryDatabase.db")
                Complete_df = pd.read_sql_query("SELECT * FROM Eagle_BadGSRInventoryDatabase ORDER BY `Unit_SN` ASC ;", conn)
                data = pd.DataFrame(Complete_df)
                data ['DuplicatedEntries']=data.sort_values(by =['Date']).duplicated(['Unit_SN','DeviceType'],keep='last')
                
                data_Duplicated = data.loc[data.DuplicatedEntries == True, 'BatchNumber': 'FaultFound']
                data_View_Duplicated = data_Duplicated.reset_index(drop=True)
                data_View_Duplicated.to_sql('Eagle_BadGSRInventoryDatabase_ACCUMULATED_DUPLICATED',conn, if_exists="append", index=False)
                
                data = data.loc[data.DuplicatedEntries == False, 'BatchNumber': 'FaultFound']
                data = data.reset_index(drop=True)
                self.cur=conn.cursor()
                data.to_sql('Eagle_BadGSRInventoryDatabase',conn, if_exists="replace", index=False)
                for each_rec in range(len(data)):
                    tree.insert("", tk.END, values=list(data.loc[each_rec]))
                tkinter.messagebox.showinfo("Update Master DB Complete","Old Duplicated Unit CaseSrNo Entries Are Removed And Replaced By Newer One")
                self.txtBatchNumber['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_BatchNumber())))
                self.txtJobName['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_JobName())))
                self.txtCrewNumber['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_CrewNumber())))
                self.txtLocation['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_Location())))
                self.txtDate['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_Date())))
                self.txtUnit_SN['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_Unit_SN())))
                self.txtDeviceType['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_DeviceType())))
                self.txtOpened['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_Opened())))
                self.txtFaultFound['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_FaultFound())))  
                conn.commit()
                conn.close()
                TotalInvCount()
                DuplicatedInvCount()

        def TotalInvCount():
            self.txtTotalEntries.delete(0,END)
            conn = sqlite3.connect("Eagle_BadGSRInventoryDatabase.db")
            Inv_df = pd.read_sql_query("select * from Eagle_BadGSRInventoryDatabase ;", conn)
            Inv_count_data = pd.DataFrame(Inv_df)
            Total_count = Inv_count_data['Unit_SN'].count()
            self.txtTotalEntries.insert(tk.END,Total_count)
            conn.commit()
            conn.close()


        def DuplicatedInvCount():
            self.txtDuplicatedEntries.delete(0,END)
            conn = sqlite3.connect("Eagle_BadGSRInventoryDatabase.db")
            Inv_df = pd.read_sql_query("select * from Eagle_BadGSRInventoryDatabase_ACCUMULATED_DUPLICATED ;", conn)
            Inv_count_data = pd.DataFrame(Inv_df)
            DuplicatedEntries = Inv_count_data['Unit_SN'].count()
            self.txtDuplicatedEntries.insert(tk.END,DuplicatedEntries)
            conn.commit()
            conn.close()

        def UpdateMASTERBeforePopulateDB():
            conn = sqlite3.connect("Eagle_BadGSRInventoryDatabase.db")
            Complete_df = pd.read_sql_query("SELECT * FROM Eagle_BadGSRInventoryDatabase ORDER BY `Unit_SN` ASC ;", conn)
            data = pd.DataFrame(Complete_df)
            data ['DuplicatedEntries']=data.sort_values(by =['Date']).duplicated(['Unit_SN','DeviceType'],keep='last')

            data_Duplicated = data.loc[data.DuplicatedEntries == True, 'BatchNumber': 'FaultFound']
            data_View_Duplicated = data_Duplicated.reset_index(drop=True)
            data_View_Duplicated.to_sql('Eagle_BadGSRInventoryDatabase_ACCUMULATED_DUPLICATED',conn, if_exists="append", index=False)
            

            data = data.loc[data.DuplicatedEntries == False, 'BatchNumber': 'FaultFound']
            data = data.reset_index(drop=True)
            self.cur=conn.cursor()
            data.to_sql('Eagle_BadGSRInventoryDatabase',conn, if_exists="replace", index=False)                
            conn.commit()
            conn.close()
            

            
        def ViewDuplicated():
            tree.delete(*tree.get_children())
            self.txtDuplicatedEntries.delete(0,END)
            conn = sqlite3.connect("Eagle_BadGSRInventoryDatabase.db")
            Duplicated_df = pd.read_sql_query("SELECT * FROM Eagle_BadGSRInventoryDatabase_ACCUMULATED_DUPLICATED ORDER BY `Unit_SN` ASC ;", conn)
            data = pd.DataFrame(Duplicated_df)
            DuplicatedEntries = len(data)
            for each_rec in range(len(data)):
                tree.insert("", tk.END, values=list(data.loc[each_rec]))
            conn.commit()
            conn.close()
            self.txtDuplicatedEntries.insert(tk.END,DuplicatedEntries)
                    
        def iExit():
            iExit= tkinter.messagebox.askyesno("Eagle Bad BoxInventory Management System", "Confirm if you want to exit")
            if iExit >0:
                global root
                root.destroy()
                return

        def ClearData():
            self.txtBatchNumber.delete(0,END)
            self.txtJobName.delete(0,END)
            self.txtCrewNumber.delete(0,END)
            self.txtLocation.delete(0,END)
            self.txtDate.delete(0,END)            
            self.txtUnit_SN.delete(0,END)
            self.txtDeviceType.delete(0,END)
            self.txtOpened.delete(0,END)
            self.txtFaultFound.delete(0,END)            
            self.txtKeySearch.delete(0,END)

        def Reset_TotalDBCount():
            self.txtTotalEntries.delete(0,END)
            
        def Reset_TotalDuplicateCount():
            self.txtDuplicatedEntries.delete(0,END)
        
        def AddData():
            if(len(BatchNumber.get())!=0) & (len(Unit_SN.get())!=0) & (len(JobName.get())!=0) & (len(Date.get())!=0)& (len(CrewNumber.get())!=0) & (len(Location.get())!=0) & (len(DeviceType.get())!=0) & (len(Opened.get())!=0) & (len(FaultFound.get())!=0):                
                Eagle_BadGSRInventoryDatabase_BackEnd.addInvRec(BatchNumber.get(), JobName.get(), CrewNumber.get(), Location.get(),Date.get(), Unit_SN.get(), DeviceType.get(), Opened.get(), FaultFound.get())
                tree.delete(*tree.get_children())
                tree.insert("", tk.END,values=(BatchNumber.get(), JobName.get(), CrewNumber.get(), Location.get(),Date.get(), Unit_SN.get(), DeviceType.get(), Opened.get(), FaultFound.get()))                                
            else:
                tkinter.messagebox.showerror("Add Error","All Entries Must Be Entered, Any Field can not be empty")
                        
            self.txtBatchNumber['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_BatchNumber())))
            self.txtJobName['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_JobName())))
            self.txtCrewNumber['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_CrewNumber())))
            self.txtLocation['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_Location())))
            self.txtDate['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_Date())))
            self.txtUnit_SN['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_Unit_SN())))
            self.txtDeviceType['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_DeviceType())))
            self.txtOpened['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_Opened())))
            self.txtFaultFound['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_FaultFound())))            
            TotalInvCount()
            self.txtUnit_SN.delete(0,END)
            self.txtFaultFound.delete(0,END)

        
                    
        def ViewMasterDB():
            UpdateMASTERBeforePopulateDB()
            tree.delete(*tree.get_children())
            for row in Eagle_BadGSRInventoryDatabase_BackEnd.viewData():
                tree.insert("", tk.END, values=row)
            self.txtBatchNumber['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_BatchNumber())))
            self.txtJobName['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_JobName())))
            self.txtCrewNumber['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_CrewNumber())))
            self.txtLocation['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_Location())))
            self.txtDate['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_Date())))
            self.txtUnit_SN['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_Unit_SN())))
            self.txtDeviceType['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_DeviceType())))
            self.txtOpened['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_Opened())))
            self.txtFaultFound['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_FaultFound())))
            TotalInvCount()
            DuplicatedInvCount()

        def searchDatabase():
            tree.delete(*tree.get_children())
            for row in Eagle_BadGSRInventoryDatabase_BackEnd.searchData(BatchNumber.get(), JobName.get(), CrewNumber.get(), Location.get(),Date.get(), Unit_SN.get(), DeviceType.get(), Opened.get(), FaultFound.get()):
                tree.insert("", tk.END, values=row)
            

        def InventoryRec(event):
            for nm in tree.selection():
                sd = tree.item(nm, 'values')
                self.txtBatchNumber.delete(0,END)
                self.txtBatchNumber.insert(tk.END,sd[0])                
                self.txtJobName.delete(0,END)
                self.txtJobName.insert(tk.END,sd[1])
                self.txtCrewNumber.delete(0,END)
                self.txtCrewNumber.insert(tk.END,sd[2])
                self.txtLocation.delete(0,END)
                self.txtLocation.insert(tk.END,sd[3])
                self.txtDate.delete(0,END)
                self.txtDate.insert(tk.END,sd[4])                
                self.txtUnit_SN.delete(0,END)
                self.txtUnit_SN.insert(tk.END,sd[5])
                self.txtDeviceType.delete(0,END)
                self.txtDeviceType.insert(tk.END,sd[6])
                self.txtOpened.delete(0,END)
                self.txtOpened.insert(tk.END,sd[7])
                self.txtFaultFound.delete(0,END)
                self.txtFaultFound.insert(tk.END,sd[8])
     

        def DeleteData():
            iDelete = tkinter.messagebox.askyesno("Delete Entry From Database", "Confirm if you want to Delete")
            if iDelete >0:
                conn = sqlite3.connect("Eagle_BadGSRInventoryDatabase.db")
                cur = conn.cursor()
                if(len(BatchNumber.get())!=0) & (len(Unit_SN.get())!=0):
                    for selected_item in tree.selection():
                        cur.execute("DELETE FROM Eagle_BadGSRInventoryDatabase WHERE BatchNumber =? AND JobName=? AND \
                                CrewNumber =? AND Location =? AND Date =? AND Unit_SN =? AND DeviceType =? AND Opened =? AND FaultFound =? ",\
                                (tree.set(selected_item, '#1'), tree.set(selected_item, '#2'),tree.set(selected_item, '#3'),\
                                 tree.set(selected_item, '#4'),tree.set(selected_item, '#5'),tree.set(selected_item, '#6'),\
                                 tree.set(selected_item, '#7'),tree.set(selected_item, '#8'),tree.set(selected_item, '#9'),))
                        conn.commit()
                        tree.delete(selected_item)
                    conn.commit()
                    conn.close()
                ClearData()
                ViewMasterDB()
                TotalInvCount()
                return


        def UpdateSlectedBatchNumber():
            iUpdateSlectedBatchNumber = tkinter.messagebox.askyesno("Update BatchNumber in Database", "Confirm if you want to Update BatchNumber")
            if iUpdateSlectedBatchNumber >0:
                conn = sqlite3.connect("Eagle_BadGSRInventoryDatabase.db")
                cur = conn.cursor()
                application_window = self.root
                BatchNumber_update = simpledialog.askstring("Input Updated BatchNumber", "What is your updated BatchNumber?",
                                parent=application_window)
                if BatchNumber_update is not None:
                    for selected_item in tree.selection():
                        cur.execute("UPDATE Eagle_BadGSRInventoryDatabase SET BatchNumber =? WHERE JobName=? AND Date =? AND Unit_SN =? AND FaultFound =?",
                                    (BatchNumber_update, tree.set(selected_item, '#2'), tree.set(selected_item, '#5'),
                                     tree.set(selected_item, '#6'), tree.set(selected_item, '#9')))
                        conn.commit()                        
                    conn.commit()
                    conn.close()
                    ViewMasterDB()
                else:
                    tkinter.messagebox.showinfo("Update Error","Please Input Updated Batch Number") 

                return

        def UpdateSlectedFaultFound():
            iUpdateSlectedFaultFound = tkinter.messagebox.askyesno("Update FaultFound in Database", "Confirm if you want to Update FaultFound")
            if iUpdateSlectedFaultFound >0:
                conn = sqlite3.connect("Eagle_BadGSRInventoryDatabase.db")
                cur = conn.cursor()
                application_window = self.root
                FaultFound_update = simpledialog.askstring("Input Updated FaultFound", "What is your updated FaultFound?",
                                parent=application_window)
                if FaultFound_update is not None:
                    for selected_item in tree.selection():
                        cur.execute("UPDATE Eagle_BadGSRInventoryDatabase SET FaultFound =? WHERE BatchNumber =? AND JobName =? AND Date =? AND Unit_SN =?",
                                    (FaultFound_update, tree.set(selected_item, '#1'), tree.set(selected_item, '#2'),
                                     tree.set(selected_item, '#5'), tree.set(selected_item, '#6')))
                                     
                        conn.commit()                        
                    conn.commit()
                    conn.close()
                    ViewMasterDB()
                else:
                    tkinter.messagebox.showinfo("Update Error","Please Input Updated FaultFound") 
                return

        def UpdateSlectedJobName():
            iUpdateSlectedJobName = tkinter.messagebox.askyesno("Update JobName in Database", "Confirm if you want to Update JobName")
            if iUpdateSlectedJobName >0:
                conn = sqlite3.connect("Eagle_BadGSRInventoryDatabase.db")
                cur = conn.cursor()
                application_window = self.root
                JobName_update = simpledialog.askstring("Input Updated JobName", "What is your updated JobName?",
                                parent=application_window)
                if JobName_update is not None:
                    for selected_item in tree.selection():
                        cur.execute("UPDATE Eagle_BadGSRInventoryDatabase SET JobName =? WHERE BatchNumber =? AND Date =? AND Unit_SN =? AND FaultFound =?",
                                    (JobName_update, tree.set(selected_item, '#1'), tree.set(selected_item, '#5'),
                                     tree.set(selected_item, '#6'), tree.set(selected_item, '#9')))
                        conn.commit()                        
                    conn.commit()
                    conn.close()
                    ViewMasterDB()
                else:
                    tkinter.messagebox.showinfo("Update Error","Please Input Updated JobName") 
                return

        def GenerateTransmittalOut():
            iGenerateTransmittalOut = tkinter.messagebox.askyesno("Generate Transmittal Out", "Confirm if you want to Generate Transmittal")
            if iGenerateTransmittalOut >0:
                application_window = self.root
                BatchNameEntry = simpledialog.askstring("Input Batch Number For Transmittal Out", "Please Input Your Batch Number To Generate Transmittal Output?",
                                parent=application_window)
                
                if BatchNameEntry is not None:
                    ## Connect SQL and Seach Query
                    list_of_values = BatchNameEntry.split(",")
                    conn = sqlite3.connect("Eagle_BadGSRInventoryDatabase.db")
                    Complete_df = pd.read_sql_query("SELECT * FROM Eagle_BadGSRInventoryDatabase;", conn)
                    data = pd.DataFrame(Complete_df)                    
                    data = data[data['BatchNumber'].isin(list_of_values)]
                    data = data.sort_values(by =['BatchNumber'])
                    data = data.reset_index(drop=True)
                    data = pd.DataFrame(data)
                    
                    ## Tree View
                    window = Tk()
                    window.title("Generated Transmittal Output View")
                    window.config(bg="ghost white")
                    width = 1250
                    height = 880
                    screen_width = window.winfo_screenwidth()
                    screen_height = window.winfo_screenheight()
                    x = (screen_width/2) - (width/2)
                    y = (screen_height/2) - (height/2)
                    window.geometry("%dx%d+%d+%d" % (width, height, x, y))
                    window.resizable(0, 0)                    
                    TableMargin = Frame(window, bd = 2, pady= 6)
                    TableMargin.pack(side=TOP)
                    TableMargin.pack(side=LEFT)
                    scrollbarx = Scrollbar(TableMargin, orient=HORIZONTAL)
                    scrollbary = Scrollbar(TableMargin, orient=VERTICAL)
                    tree = ttk.Treeview(TableMargin, column=("column1", "column2", "column3", "column4", "column5", "column6", "column7", "column8", "column9"),
                                                height=34, show='headings')
                    scrollbary.config(command=tree.yview)
                    scrollbary.pack(side=RIGHT, fill=Y)
                    scrollbarx.config(command=tree.xview)
                    scrollbarx.pack(side=BOTTOM, fill=X)
                    tree.heading("#1", text="BatchNumber", anchor=W)
                    tree.heading("#2", text="JobName", anchor=W)
                    tree.heading("#3", text="CrewNumber", anchor=W)
                    tree.heading("#4", text="Location", anchor=W)
                    tree.heading("#5", text="Date", anchor=W)            
                    tree.heading("#6", text="Unit_SN", anchor=W)
                    tree.heading("#7", text="DeviceType" ,anchor=W)
                    tree.heading("#8", text="Opened", anchor=W)
                    tree.heading("#9", text="FaultFound", anchor=W)                    
                    tree.column('#1', stretch=NO, minwidth=0, width=90)            
                    tree.column('#2', stretch=NO, minwidth=0, width=100)
                    tree.column('#3', stretch=NO, minwidth=0, width=90)
                    tree.column('#4', stretch=NO, minwidth=0, width=80)
                    tree.column('#5', stretch=NO, minwidth=0, width=70)
                    tree.column('#6', stretch=NO, minwidth=0, width=80)
                    tree.column('#7', stretch=NO, minwidth=0, width=80)
                    tree.column('#8', stretch=NO, minwidth=0, width=80)
                    tree.column('#9', stretch=NO, minwidth=0, width=100)                    
                    tree.pack()
                    ## Label and Entry
                    TitFrame = Frame(window, bd = 2, padx= 2, pady= 2, bg = "#006dcc", relief = RIDGE)
                    TitFrame.pack(side = RIGHT)
                    TitFrame.pack(side = TOP)
                    self.lblTit = Label(TitFrame, font=('aerial', 10, 'bold'),
                                        text="Eagle Canada Equipment Transmittal Out \n 6806 Railway Street SE \n Calgary, AB T2H 3A8\n Ph: (403) 263-7770",
                                        bg="#006dcc")
                    self.lblTit.grid()

                    L1 = Label(window, text = "A: Count in Transmittal :", font=("arial", 10,'bold'),bg = "ghost white").place(x=6,y=8)
                    self.txtTotalTransmittalEntry  = Entry(window, font=('aerial', 12, 'bold'),textvariable = IntVar(), width = 11, bd=2)
                    self.txtTotalTransmittalEntry.place(x=180,y=8)

                    L2 = Label(window, text = "B: Transmittal Out Date:", font=("arial", 10,'bold'),bg = "ghost white").place(x=6,y=48)
                    self.txtTransmittalDate  = Entry(window, font=('aerial', 12, 'bold'),textvariable = StringVar(window,value=Default_Date_today), width = 11, bd=2)
                    self.txtTransmittalDate.place(x=180,y=48)

                    L3 = Label(window, text = "C: Transmittal Number :", font=("arial", 10,'bold'),bg = "ghost white").place(x=390,y=8)
                    self.txtTransmittalBatch  = Entry(window, font=('aerial', 12, 'bold'),textvariable = IntVar(), width = 15, bd=2)
                    self.txtTransmittalBatch.place(x=559,y=8)

                    Sending_Reason = ["For Repair - Bad Equipment", "For Crew - Production", "Job End - Transfer To Calgary Shop", "Crew To Crew Transfer - Production", "Crew To Crew Transfer - Repair"]
                    L14 = Label(window, text = "D: Transmittal Reason :", font=("arial", 10,'bold'),bg = "ghost white").place(x=390,y=48)
                    self.txtReasonSending = ttk.Combobox(window, font=('aerial', 10, 'bold'),  width = 31, values= Sending_Reason)
                    self.txtReasonSending.current(0)
                    self.txtReasonSending.place(x=559,y=48)

                    L4 = Label(window, text = "E: Transmittal Out Job/Program Information :", font=("arial", 12,'bold'),bg = "ghost white").place(x=800,y=100)

                    L5 = Label(window, text = "1: Job/Program Name :", font=("arial", 10,'bold'),bg = "ghost white").place(x=800,y=130)
                    JobName = StringVar()
                    self.txtProjectName = ttk.Combobox(window, font=('aerial', 10, 'bold'), textvariable = JobName, width = 31)
                    self.txtProjectName.place(x=980,y=130)
                    self.txtProjectName['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_ProjectName())))
                    self.txtProjectName.current(0)

                    L6 = Label(window, text = "2: Job/Program Location :", font=("arial", 10,'bold'),bg = "ghost white").place(x=800,y=170)
                    Location = StringVar()
                    self.txtProjectLocation  = ttk.Combobox(window, font=('aerial', 10, 'bold'), textvariable = Location, width = 31)
                    self.txtProjectLocation.place(x=980,y=170)
                    self.txtProjectLocation['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_ProjectLocation())))
                    self.txtProjectLocation.current(0)

                    L7 = Label(window, text = "3: Crew Number :", font=("arial", 10,'bold'),bg = "ghost white").place(x=800,y=210)
                    Crew_Number = StringVar()
                    self.txtCrew_Number = ttk.Combobox(window, font=('aerial', 10, 'bold'), textvariable = Crew_Number, width = 31)
                    self.txtCrew_Number.place(x=980,y=210)
                    self.txtCrew_Number['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_ProjectCrewNumber())))
                    self.txtCrew_Number.current(0)

                    L8 = Label(window, text = "4: Job/Program Number :", font=("arial", 10,'bold'),bg = "ghost white").place(x=800,y=250)
                    self.txtProgramNumber  = Entry(window, font=('aerial', 10, 'bold'),textvariable = StringVar(), width = 33, bd=2)
                    self.txtProgramNumber.place(x=980,y=250)

                    L9 = Label(window, text = "5: Crew Manager Name :", font=("arial", 10,'bold'),bg = "ghost white").place(x=800,y=290)
                    self.txtCrewManager  = Entry(window, font=('aerial', 10, 'bold'),textvariable = StringVar(), width = 33, bd=2)
                    self.txtCrewManager.place(x=980,y=290)


                    L10 = Label(window, text = "F: Transmittal Receiver Information :", font=("arial", 12,'bold'),bg = "ghost white").place(x=800,y=340)

                    L11 = Label(window, text = "1: Receiver Name :", font=("arial", 10,'bold'),bg = "ghost white").place(x=800,y=380)
                    self.txtReceiverName  = Entry(window, font=('aerial', 10, 'bold'),textvariable = StringVar(), width = 33, bd=2)
                    self.txtReceiverName.place(x=980,y=380)

                    L12 = Label(window, text = "2: Receiver Address :", font=("arial", 10,'bold'),bg = "ghost white").place(x=800,y=420)
                    ReceiverAddress = StringVar()
                    self.txtReceiverAddress  = ttk.Combobox(window, font=('aerial', 10, 'bold'),textvariable = ReceiverAddress, width = 31)
                    self.txtReceiverAddress.place(x=980,y=420)
                    self.txtReceiverAddress['values'] = list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_ReceiverLocation()))
                    

                    Equipment_Type_List = ["Recording", "Source", "Vibroseis", "Survey", "IT", "HSE"]                    
                    L13 = Label(window, text = "3: Equipment Type :", font=("arial", 10,'bold'),bg = "ghost white").place(x=800,y=460)
                    self.txtEquipmentType  = ttk.Combobox(window, font=('aerial', 10, 'bold'),  width = 31, values= Equipment_Type_List)
                    self.txtEquipmentType.current(0)
                    self.txtEquipmentType.place(x=980,y=460)

                    Equipment_Name_List = ["Geospace GSR/GSX", "Geospace GSI/SDR", "Inovageo Hawk", "Seismic Source", "Geospace LineViewer", "Hawk LineViewer", "ARAM Cable"]
                    L14 = Label(window, text = "4: Equipment Name :", font=("arial", 10,'bold'),bg = "ghost white").place(x=800,y=500)
                    self.txtEquipmentName  = ttk.Combobox(window, font=('aerial', 10, 'bold'),  width = 31, values= Equipment_Name_List)
                    self.txtEquipmentName.current(0)
                    self.txtEquipmentName.place(x=980,y=500)

                    L15 = Label(window, text = "5: Comments :", font=("arial", 10,'bold'),bg = "ghost white").place(x=800,y=540)
                    self.txtComments  = Entry(window, font=('aerial', 10, 'bold'),textvariable = StringVar(), width = 33, bd=2)                    
                    self.txtComments.place(x=980,y=540)

                    L16 = Label(window, text = "G: Transportation/Additional Information :", font=("arial", 12,'bold'),bg = "ghost white").place(x=800,y=590)

                    L17 = Label(window, text = "1: Shipper Name :", font=("arial", 10,'bold'),bg = "ghost white").place(x=800,y=630)
                    self.txtDriverName  = Entry(window, font=('aerial', 10, 'bold'),textvariable = StringVar(), width = 33, bd=2)
                    self.txtDriverName.place(x=980,y=630)

                    L18 = Label(window, text = "2: Transported By :", font=("arial", 10,'bold'),bg = "ghost white").place(x=800,y=670)
                    self.txtVehicalNumber  = Entry(window, font=('aerial', 10, 'bold'),textvariable = StringVar(), width = 33, bd=2)
                    self.txtVehicalNumber.place(x=980,y=670)

                    L19 = Label(window, text = "H: Eagle TDG Form Completed For Batteries? :", font=("arial", 10,'bold'),bg = "ghost white").place(x=250,y=830)
                    Answer_List = ["No, Not Required", "Yes"]                                        
                    self.txtTDG  = ttk.Combobox(window, font=('aerial', 10, 'bold'),  width = 20, values= Answer_List)
                    self.txtTDG.current(0)
                    self.txtTDG.place(x=560,y=830)

                    L20 = Label(window, text = "3: Total Weight (lbs) :", font=("arial", 10,'bold'),bg = "ghost white").place(x=800,y=710)
                    self.txtTotalWeight  = Entry(window, font=('aerial', 10, 'bold'),textvariable = StringVar(), width = 33, bd=2)
                    self.txtTotalWeight.place(x=980,y=710)

                    Owner_Name_List = ["Geospace Tech", "Eagle Canada", "Dawson Geophysical", "Seismic Source", "INOVA Geophysical", "Geo-Check", "Mitcham","Global Geo"]
                    L21 = Label(window, text = "4: Owner Name :", font=("arial", 10,'bold'),bg = "ghost white").place(x=800,y=750)
                    self.txtOwnerName  = ttk.Combobox(window, font=('aerial', 10, 'bold'),  width = 31, values= Owner_Name_List)
                    self.txtOwnerName.place(x=980,y=750)
                    self.txtOwnerName.current(0)

                    L22 = Label(window, text = "5: PO Number :", font=("arial", 10,'bold'),bg = "ghost white").place(x=800,y=790)
                    self.txtPONumber  = Entry(window, font=('aerial', 10, 'bold'),textvariable = StringVar(), width = 33, bd=2)
                    self.txtPONumber.place(x=980,y=790)

                    ## Define functions
                    
                    def iExit():
                        window.destroy()
                    
                    def ClearAll():
                        self.txtTotalTransmittalEntry.delete(0,END)
                        self.txtTransmittalDate.delete(0,END)
                        self.txtTransmittalBatch.delete(0,END)                                                
                        self.txtProjectName.delete(0,END)
                        self.txtProjectLocation.delete(0,END)
                        self.txtCrew_Number.delete(0,END)
                        self.txtProgramNumber.delete(0,END)
                        self.txtCrewManager.delete(0,END)
                        self.txtDriverName.delete(0,END)
                        self.txtVehicalNumber.delete(0,END)                        
                        self.txtReceiverName.delete(0,END)
                        self.txtReceiverAddress.delete(0,END)
                        self.txtEquipmentType.delete(0,END)
                        self.txtEquipmentName.delete(0,END)
                        self.txtComments.delete(0,END)
                        self.txtReasonSending.delete(0,END)
                        self.txtTotalWeight.delete(0,END)
                        self.txtOwnerName.delete(0,END)
                        self.txtTDG.delete(0,END)
                        self.txtPONumber.delete(0,END)
                        
                    def TransmittalRec(event):
                        for nm in tree.selection():
                            sd = tree.item(nm, 'values')

                    def DeleteSelectData():
                        iDelete = tkinter.messagebox.askyesno("Delete Entry From Transmittal", "Confirm if you want to Delete")
                        if iDelete >0:
                            self.txtTotalTransmittalEntry.delete(0,END)
                            for selected_item in tree.selection():
                                tree.delete(selected_item)
                            Total_count = len(tree.get_children())
                            self.txtTotalTransmittalEntry.insert(tk.END,Total_count)
                        return

                    def ExportTransmittal():
                        Total_Count_Export = self.txtTotalTransmittalEntry.get()
                        Total_Count = ("Transmittal Quantity: ")

                        TransmittalDate_Export = self.txtTransmittalDate.get()
                        TransmittalDate = ("Transmittal Date: ")

                        TransmittalNumber_Export = self.txtTransmittalBatch.get()
                        TransmittalNumber = ("Transmittal No : ")                        

                        ProjectName_Export = self.txtProjectName.get()
                        ProjectName = ("Job/Program Name : ")

                        ProjectLocation_Export = self.txtProjectLocation.get()
                        ProjectLocation = ("Job/Program Location : ")

                        CrewNumber_Export = self.txtCrew_Number.get()
                        CrewNumber = ("Crew Number : ")

                        ProgramNumber_Export = self.txtProgramNumber.get()
                        ProgramNumber = ("Job/Program Number : ")

                        CrewManager_Export = self.txtCrewManager.get()
                        CrewManager = ("Crew Manager Name : ")
                        
                        EquipmentType_Export = self.txtEquipmentType.get()
                        EquipmentType = ("Equipment Type : ")

                        EquipmentName_Export = self.txtEquipmentName.get()
                        EquipmentName = ("Equipment Name : ")

                        ReceiverName_Export = self.txtReceiverName.get()
                        ReceiverName = ("Receiver Name : ")

                        VehicalNumber_Export = self.txtVehicalNumber.get()
                        VehicalNumber = ("Transported by : ")

                        DriverName_Export = self.txtDriverName.get()
                        DriverName = ("Shipper Name : ")
                        
                        Comments_Export = self.txtComments.get()
                        Comments = ("E: Comments: ")

                        ReceiverAddress_Export = self.txtReceiverAddress.get()
                        ReceiverAddress = ("Receiver Address : ")

                        SendingReason_Export = self.txtReasonSending.get()

                        Owner_Export        = self.txtOwnerName.get()
                        Weight_Export       = self.txtTotalWeight.get()
                        TDG_Export          = self.txtTDG.get()
                        PO_Number_Export    = self.txtPONumber.get()
                        

                        TransmittalSummary = (['',''],
                                            [Total_Count, Total_Count_Export],
                                            [TransmittalDate,   TransmittalDate_Export],
                                            [TransmittalNumber,  TransmittalNumber_Export],['',''],
                                            ['',''],
                                            [ProjectName,  ProjectName_Export],
                                            [ProjectLocation,  ProjectLocation_Export],
                                            [CrewNumber,  CrewNumber_Export],
                                            [ProgramNumber,  ProgramNumber_Export],
                                            [CrewManager,  CrewManager_Export],
                                            [DriverName,  DriverName_Export],
                                            [VehicalNumber,  VehicalNumber_Export])

                        ReceiverSummary = (['',''],
                                           [ReceiverName, ReceiverName_Export],
                                           [EquipmentType,  EquipmentType_Export],
                                           [EquipmentName,  EquipmentName_Export],['',''],
                                           ['',''])

                        ReceiverAddress_Split = ReceiverAddress_Export.split(",")
                                                
                        dfList =[] 
                        for child in tree.get_children():
                            df = tree.item(child)["values"]
                            dfList.append(df)
                        Transmittal_DF = pd.DataFrame(dfList)
                        Transmittal_DF.rename(columns = {0:'Batch Number', 1:'Project Name', 2:'Crew Number', 3:'Project Location', 4:'Date Reported',
                                                      5: 'Unit Serial Number', 6:'Device Type', 7:'Opened (Y/N)', 8:'Fault Found'},inplace = True)
                        Transmittal_DF_SortByCaseSrNo = Transmittal_DF.sort_values(by =['Batch Number'])

                        ## Create Transmittal Front DataFrame

                        Transmittal_Front_DF = pd.DataFrame()

                        Transmittal_Front_A = (['Date:',TransmittalDate_Export],
                                           ['Shipper:', DriverName_Export],
                                           ['Receiver:',  ReceiverName_Export],
                                           ['PO Number:', PO_Number_Export ])

                        Transmittal_Front_B = (['From Location:',ProjectLocation_Export],
                                           ['To Location:', ReceiverAddress_Split[0]],
                                           ['Other To/From:',  ''],
                                           ['Transported By:',  VehicalNumber_Export])

                        Transmittal_Front_C = (['Crew:',CrewNumber_Export],
                                           ['Program Name:', ProjectName_Export],
                                           ['Program Number:',  ProgramNumber_Export],
                                           ['Crew Manager:',  CrewManager_Export])
                        
                        
                        ### Working with the Selected File                       
                        filename = tkinter.filedialog.asksaveasfilename(initialdir = "/",title = "Select file" ,
                                   defaultextension='.xlsx', filetypes = (("Excel file",".xlsx"),("Excel file",".xlsx")))
                        if filename:
                            if filename.endswith('.xlsx'):
                                with pd.ExcelWriter(filename, engine='xlsxwriter') as file:

                                    Transmittal_Front_DF.to_excel(file,sheet_name='Transmittal Front Page',index=False, startrow=6, header=True)                                    
                                    Transmittal_DF_SortByCaseSrNo.to_excel(file,sheet_name='Equipment List Transmittal',index=False, startrow=16, header=False)
                                    workbook_ListBadGSR  = file.book
                                    worksheet_Front      = file.sheets['Transmittal Front Page']
                                    worksheet_ListBadGSR = file.sheets['Equipment List Transmittal']

                                    worksheet_Front.set_margins(0.3, 0.4, 1.6, 1.1)                                    
                                    worksheet_Front.set_landscape()
                                    worksheet_Front.print_area('A1:O27')
                                    worksheet_Front.print_across()
                                    worksheet_Front.fit_to_pages(1, 1)                                    
                                    worksheet_Front.set_paper(9)
                                    worksheet_Front.set_start_page(1)
                                    worksheet_Front.hide_gridlines(0)
                                    worksheet_Front.set_page_view()
                                    
                                    worksheet_ListBadGSR.set_margins(0.3, 0.1, 1.6, 1.1)                                    
                                    worksheet_ListBadGSR.set_portrait()
                                    worksheet_ListBadGSR.print_area('A1:I44')
                                    worksheet_ListBadGSR.print_across()
                                    worksheet_ListBadGSR.fit_to_pages(1, 0)                                    
                                    worksheet_ListBadGSR.set_paper(9)
                                    worksheet_ListBadGSR.set_start_page(1)
                                    worksheet_ListBadGSR.hide_gridlines(0)
                                    worksheet_ListBadGSR.set_page_view()

                                    headerFront = '&L&G'+'&CEagle Canada Seismic Services ULC' + '\n' + '6806 Railway Street SE' + '\n' + 'Calgary, AB T2H 3A8' + '\n' +  'Ph: (403) 263-7770' +  '&R&U&24&"cambria, bold"Transmittal'
                                    header1 = '&L&G'+'&CEagle Canada Seismic Services ULC' + '\n' + '6806 Railway Street SE' + '\n' + 'Calgary, AB T2H 3A8' + '\n' +  'Ph: (403) 263-7770' +  '&R&U&18&"cambria, bold"Transmittal' +'\n' +'Bad Box Unit'
                                    worksheet_Front.set_header(headerFront,{'image_left':'eagle logo.jpg'})
                                    worksheet_ListBadGSR.set_header(header1,{'image_left':'eagle logo.jpg'})
                                    footerFront = ('&LRevised Date : &D')+ ('&CEFXX-XX-XX') + ('&RReceiver Name (Print): &Y-------------------------------------------------------------' + '\n' + '&RReceiver Signature: &Y-------------------------------------------------------------')
                                    footer1 = ('&LDate : &D') + ('&RReceiver Name (Print): &Y-----------------------------------------' + '\n' + '&RReceiver Signature: &Y-----------------------------------------')
                                    worksheet_ListBadGSR.set_footer(footer1)
                                    worksheet_Front.set_footer(footerFront)
                                    
                                    cell_format_1 = workbook_ListBadGSR.add_format({'bold': True, 'text_wrap': True, 'align': 'top', 'valign': 'top', 'border': 0})
                                    cell_format_1.set_underline(1)
                                    cell_format_2 = workbook_ListBadGSR.add_format({'bold': False, 'text_wrap': True, 'align': 'top', 'valign': 'top', 'border': 0})
                                    cell_format_3 = workbook_ListBadGSR.add_format({'bold': False, 'text_wrap': True, 'align': 'top', 'valign': 'top', 'border': 0})
                                    cell_format_3.set_font_size(15)
                                    cell_format_4 = workbook_ListBadGSR.add_format({'bold': True, 'text_wrap': True, 'align': 'bottom', 'valign': 'top', 'border': 0})
                                    cell_format_4.set_underline(1)
                                    cell_format_5 = workbook_ListBadGSR.add_format({'bold': True, 'text_wrap': True, 'align': 'center', 'valign': 'center', 'border': 0})
                                    cell_format_5.set_underline(1)
                                    cell_format_6 = workbook_ListBadGSR.add_format({'bold': False, 'text_wrap': True, 'align': 'center', 'valign': 'top', 'border': 0})
                                    cell_format_6.set_font_size(12)

                                    cell_format_7 = workbook_ListBadGSR.add_format({'bold': True, 'text_wrap': True, 'align': 'left', 'valign': 'top', 'border': 0})
                                    cell_format_7.set_font_size(13)

                                    worksheet_Front.merge_range('A1:B1', "")
                                    worksheet_Front.merge_range('A2:B2', "")
                                    worksheet_Front.merge_range('A3:B3', "")
                                    worksheet_Front.merge_range('A4:B4', "")
                                    worksheet_Front.merge_range('C1:E1', "")
                                    worksheet_Front.merge_range('C2:E2', "")
                                    worksheet_Front.merge_range('C3:E3', "")
                                    worksheet_Front.merge_range('C4:E4', "")
                                    worksheet_Front.merge_range('F1:G1', "")
                                    worksheet_Front.merge_range('F2:G2', "")
                                    worksheet_Front.merge_range('F3:G3', "")
                                    worksheet_Front.merge_range('F4:G4', "")
                                    worksheet_Front.merge_range('H1:J1', "")
                                    worksheet_Front.merge_range('H2:J2', "")
                                    worksheet_Front.merge_range('H3:J3', "")
                                    worksheet_Front.merge_range('H4:J4', "")
                                    worksheet_Front.merge_range('K1:L1', "")
                                    worksheet_Front.merge_range('K2:L2', "")
                                    worksheet_Front.merge_range('K3:L3', "")
                                    worksheet_Front.merge_range('K4:L4', "")
                                    worksheet_Front.merge_range('M1:O1', "")
                                    worksheet_Front.merge_range('M2:O2', "")
                                    worksheet_Front.merge_range('M3:O3', "")
                                    worksheet_Front.merge_range('M4:O4', "")
                                    worksheet_Front.merge_range('A26:H27', "")

                                    worksheet_ListBadGSR.merge_range('A2:B2', "")
                                    worksheet_ListBadGSR.merge_range('A3:B3', "")
                                    worksheet_ListBadGSR.merge_range('A4:B4', "")
                                    worksheet_ListBadGSR.merge_range('A5:I5', "")                                    
                                    worksheet_ListBadGSR.merge_range('A7:B7', "")
                                    worksheet_ListBadGSR.merge_range('A8:B8', "")
                                    worksheet_ListBadGSR.merge_range('A9:B9', "")
                                    worksheet_ListBadGSR.merge_range('A10:B10', "")
                                    worksheet_ListBadGSR.merge_range('A11:B11', "")
                                    worksheet_ListBadGSR.merge_range('A12:B12', "")
                                    worksheet_ListBadGSR.merge_range('A13:B13', "")                                    
                                    worksheet_ListBadGSR.merge_range('C2:E2', "")
                                    worksheet_ListBadGSR.merge_range('C3:E3', "")
                                    worksheet_ListBadGSR.merge_range('C4:E4', "")
                                    worksheet_ListBadGSR.merge_range('C7:E7', "")
                                    worksheet_ListBadGSR.merge_range('C8:E8', "")
                                    worksheet_ListBadGSR.merge_range('C9:E9', "")
                                    worksheet_ListBadGSR.merge_range('C10:E10', "")
                                    worksheet_ListBadGSR.merge_range('C11:E11', "")
                                    worksheet_ListBadGSR.merge_range('C12:E12', "")
                                    worksheet_ListBadGSR.merge_range('C13:E13', "")
                                    worksheet_ListBadGSR.merge_range('F2:G2', "")
                                    worksheet_ListBadGSR.merge_range('F3:G3', "")
                                    worksheet_ListBadGSR.merge_range('F4:G4', "")                                    
                                    worksheet_ListBadGSR.merge_range('H2:I2', "")
                                    worksheet_ListBadGSR.merge_range('H3:I3', "")
                                    worksheet_ListBadGSR.merge_range('H4:I4', "")
                                    worksheet_ListBadGSR.merge_range('F7:I7', "")
                                    worksheet_ListBadGSR.merge_range('F8:I8', "")
                                    worksheet_ListBadGSR.merge_range('F9:I9', "")
                                    worksheet_ListBadGSR.merge_range('F10:I10', "")
                                                                        
                                    row_TransmittalSummary = 0
                                    col_TransmittalSummary = 0
                                    for item, values in (TransmittalSummary):
                                        worksheet_ListBadGSR.write(row_TransmittalSummary, col_TransmittalSummary,     item)
                                        worksheet_ListBadGSR.write(row_TransmittalSummary, col_TransmittalSummary+2, values)
                                        row_TransmittalSummary += 1
                                    
                                    row_TransmittalReceiver = 0
                                    col_TransmittalReceiver = 5
                                    for item, values in (ReceiverSummary):
                                        worksheet_ListBadGSR.write(row_TransmittalReceiver, col_TransmittalReceiver,     item)
                                        worksheet_ListBadGSR.write(row_TransmittalReceiver, col_TransmittalReceiver+2, values)
                                        row_TransmittalReceiver += 1

                                    row_ReceiverAddress = 6
                                    col_ReceiverAddress = 5
                                    for item in ReceiverAddress_Split:
                                        worksheet_ListBadGSR.write(row_ReceiverAddress, col_ReceiverAddress, item)
                                        row_ReceiverAddress += 1

                                    row_TransmittalFrontA = 0
                                    col_TransmittalFrontA = 0
                                    for item, values in (Transmittal_Front_A):
                                        worksheet_Front.write(row_TransmittalFrontA, col_TransmittalFrontA,     item)
                                        worksheet_Front.write(row_TransmittalFrontA, col_TransmittalFrontA + 2, values)
                                        row_TransmittalFrontA += 1

                                    row_TransmittalFrontB = 0
                                    col_TransmittalFrontB = 5
                                    for item, values in (Transmittal_Front_B):
                                        worksheet_Front.write(row_TransmittalFrontB, col_TransmittalFrontB,     item)
                                        worksheet_Front.write(row_TransmittalFrontB, col_TransmittalFrontB + 2, values)
                                        row_TransmittalFrontB += 1

                                    row_TransmittalFrontC = 0
                                    col_TransmittalFrontC = 10
                                    for item, values in (Transmittal_Front_C):
                                        worksheet_Front.write(row_TransmittalFrontC, col_TransmittalFrontC,     item)
                                        worksheet_Front.write(row_TransmittalFrontC, col_TransmittalFrontC + 2, values)
                                        row_TransmittalFrontC += 1

                                    cell_format_Centre = workbook_ListBadGSR.add_format()
                                    cell_format_Left = workbook_ListBadGSR.add_format()
                                    cell_format_Centre.set_align('center')
                                    cell_format_Left.set_align('left')
                                    worksheet_ListBadGSR.set_column('A:A',8, cell_format_Left)
                                    worksheet_ListBadGSR.set_column('B:B', 10, cell_format_Left)
                                    worksheet_ListBadGSR.set_column('C:C', 8, cell_format_Left)
                                    worksheet_ListBadGSR.set_column('D:D', 12, cell_format_Left)
                                    worksheet_ListBadGSR.set_column('E:E', 11, cell_format_Left)
                                    worksheet_ListBadGSR.set_column('F:F', 10, cell_format_Left)
                                    worksheet_ListBadGSR.set_column('G:G', 7, cell_format_Left)
                                    worksheet_ListBadGSR.set_column('H:H', 8, cell_format_Left)
                                    worksheet_ListBadGSR.set_column('I:I', 16, cell_format_Left)
                                    header_format_ListBadGSR = workbook_ListBadGSR.add_format({
                                                    'bold': True,
                                                    'text_wrap': True,
                                                    'valign': 'top',
                                                    'fg_color': '#808080',
                                                    'border': 2})
                                    header_format_ListBadGSR.set_align('center')
                                    worksheet_ListBadGSR.merge_range('A1:E1', "A: Transmittal Summary:", cell_format_1)
                                    worksheet_ListBadGSR.merge_range('A6:E6', "B: Transmittal Out Information:", cell_format_1)
                                    worksheet_ListBadGSR.merge_range('F1:I1', "C: Receiving Information:", cell_format_1)
                                    worksheet_ListBadGSR.merge_range('F6:I6', "D: Receiving Location:", cell_format_1)                                    
                                    worksheet_ListBadGSR.merge_range('F11:I11', "E: Reason For Transmittal:", cell_format_1)                                    
                                    worksheet_Front.merge_range('A5:G6', " Eagle TDG Form Completed For Batteries (YES / NO, Not Required)? : ", cell_format_4)
                                    worksheet_Front.merge_range('H5:O6', TDG_Export , cell_format_2)

                                    worksheet_Front.merge_range('A7:B7', "Category", cell_format_5)
                                    worksheet_Front.merge_range('A8:B25', EquipmentType_Export , cell_format_6)
                                    
                                    worksheet_Front.merge_range('C7:D7', "Item", cell_format_5)
                                    worksheet_Front.merge_range('C8:D25', EquipmentName_Export , cell_format_6)
                                    
                                    worksheet_Front.merge_range('E7:F7', "Serial #/Unit #", cell_format_5)
                                    worksheet_Front.merge_range('E8:F25', 'See Sheet Equipment List Transmittal', cell_format_6)
                                    
                                    worksheet_Front.merge_range('G7:H7', "Owner", cell_format_5)
                                    worksheet_Front.merge_range('G8:H25', Owner_Export , cell_format_6)
                                    
                                    worksheet_Front.write('I7', "Quantity", cell_format_5)
                                    worksheet_Front.merge_range('I8:I25', Total_Count_Export , cell_format_6)
                                    
                                    worksheet_Front.merge_range('J7:K7', "Weight (If Req'd)", cell_format_5)
                                    worksheet_Front.merge_range('J8:K25', "", cell_format_6)
                                    
                                    worksheet_Front.merge_range('L7:M7', "Total Weight (lbs)", cell_format_5)
                                    worksheet_Front.merge_range('L8:M25', Weight_Export , cell_format_6)
                                    
                                    worksheet_Front.merge_range('N7:O7', "Comments", cell_format_5)
                                    worksheet_Front.merge_range('N8:O25', Comments_Export , cell_format_6)

                                    worksheet_Front.merge_range('I26:K27', " Total Overall Weight (lbs) : ", cell_format_7)
                                    worksheet_Front.merge_range('L26:O27', "", cell_format_7)
                                    
                                    worksheet_ListBadGSR.merge_range('A14:B15', Comments, cell_format_1)
                                    worksheet_ListBadGSR.merge_range('C14:I15', Comments_Export, cell_format_2)
                                    worksheet_ListBadGSR.merge_range('F12:I13', SendingReason_Export, cell_format_3)                                    
                                    for col_num, value in enumerate(Transmittal_DF_SortByCaseSrNo.columns.values):
                                        worksheet_ListBadGSR.write(15, col_num, value, header_format_ListBadGSR)
                                file.close
                                tkinter.messagebox.showinfo("Transmittal Export"," Transmittal Out Report Saved as Excel")
                        tree.delete(*tree.get_children())
                        ClearAll()
                        iExit()
                        

                    ##----------------- Tree View Select Event------------
        
                    tree.bind('<<TreeviewSelect>>',TransmittalRec)
                    

                    ## Command Button
                    btnExitData = Button(window, text="Exit", font=('aerial', 10, 'bold'), height =1, width=8, bd=4, command = iExit)
                    btnExitData.place(x=4,y=820)

                    btnDeleteSelected = Button(window, text="Delete Selected", font=('aerial', 10, 'bold'), height =1, width=14, bd=4, command = DeleteSelectData)
                    btnDeleteSelected.place(x=90,y=820)

                    btnExportTransmittalOut = Button(window, text="Export Transmittal Out\n As Excel", font=('aerial', 10, 'bold'), height =2, width=19, bd=4, command = ExportTransmittal)
                    btnExportTransmittalOut.place(x=1080,y=830)
                    
                    ## Populating in TreeView Data                     
                    TotalTransmittalEntry = len(data)                 
                    self.txtTotalTransmittalEntry.insert(tk.END,TotalTransmittalEntry)
                    self.txtTransmittalBatch.insert(tk.END,BatchNameEntry)                    
                    for each_rec in range(len(data)):
                        tree.insert("", tk.END, values=list(data.loc[each_rec]))
                    
                    conn.commit()
                    conn.close()

                else:
                    tkinter.messagebox.showinfo("Generate Transmittal Message","Please Input Batch Number to Generate Transmittal") 
                return
            

        

        def ImportBatchScannedFile():
            window = Tk()
            window.title("Import Batch Scanned File Viewer")
            window.config(bg="ghost white")
            width = 1000
            height = 540
            screen_width = window.winfo_screenwidth()
            screen_height = window.winfo_screenheight()
            x = (screen_width/2) - (width/2)
            y = (screen_height/2) - (height/2)
            window.geometry("%dx%d+%d+%d" % (width, height, x, y))
            window.resizable(0, 0)
            TableMargin = Frame(window)
            TableMargin.pack(side=TOP)
            TableMargin.pack(side=LEFT)
            scrollbarx = Scrollbar(TableMargin, orient=HORIZONTAL)
            scrollbary = Scrollbar(TableMargin, orient=VERTICAL)
            tree = ttk.Treeview(TableMargin, column=("column1", "column2", "column3", "column4", "column5", "column6", "column7", "column8", "column9", "column10"),
                                        height=25, show='headings')
            scrollbary.config(command=tree.yview)
            scrollbary.pack(side=RIGHT, fill=Y)
            scrollbarx.config(command=tree.xview)
            scrollbarx.pack(side=BOTTOM, fill=X)
            tree.heading("#1", text="BatchNumber", anchor=W)
            tree.heading("#2", text="JobName", anchor=W)
            tree.heading("#3", text="CrewNumber", anchor=W)
            tree.heading("#4", text="Location", anchor=W)
            tree.heading("#5", text="Date", anchor=W)            
            tree.heading("#6", text="Unit_SN", anchor=W)
            tree.heading("#7", text="DeviceType" ,anchor=W)
            tree.heading("#8", text="Opened", anchor=W)
            tree.heading("#9", text="FaultFound", anchor=W)
            tree.heading("#10", text="Duplicated", anchor=W)
            tree.column('#1', stretch=NO, minwidth=0, width=90)            
            tree.column('#2', stretch=NO, minwidth=0, width=100)
            tree.column('#3', stretch=NO, minwidth=0, width=90)
            tree.column('#4', stretch=NO, minwidth=0, width=80)
            tree.column('#5', stretch=NO, minwidth=0, width=70)
            tree.column('#6', stretch=NO, minwidth=0, width=80)
            tree.column('#7', stretch=NO, minwidth=0, width=80)
            tree.column('#8', stretch=NO, minwidth=0, width=80)
            tree.column('#9', stretch=NO, minwidth=0, width=100)
            tree.column('#10', stretch=NO, minwidth=0, width=80)
            tree.pack()

            ##### Defining Functions For Batch Scaned Files

            def ImportBatchScannedFile():                
                name = askopenfilename(filetypes=[('Excel File', ('*.xlsx', '*.XLSX'))])
                if name:
                    extensions = ['.xlsx', '.XLSX']
                    if name.endswith(tuple(extensions)):
                        wb   =  openpyxl.load_workbook(name)
                        Sheet_Names = wb.sheetnames
                        sheetN = 0
                        dfList =[]
                        for sheetN in range(len(Sheet_Names)):
                            sheet = wb[Sheet_Names[sheetN]]
                            Batch_Number = sheet.cell(row=3, column=4)
                            Batch_Number = (Batch_Number.value)
                            Crew_Number = sheet.cell(row=5, column=4)
                            Crew_Number = (Crew_Number.value)
                            Date = sheet.cell(row=7, column=4)
                            Date = (Date.value)
                            ProjectName = sheet.cell(row=9, column=4)
                            ProjectName = (ProjectName.value)
                            ProjectLocation = sheet.cell(row=11, column=4)
                            ProjectLocation = (ProjectLocation.value)
                            
                            df = pd.read_excel(name, sheet_name=sheetN, header = None, skiprows = 17)
                            df.rename(columns = {0:'SL', 1:'Unit_SN', 2:'DeviceType', 3:'FaultFound', 4:'Opened'},inplace = True)
                            data = pd.DataFrame(df)
                            data["BatchNumber"]     = data.shape[0]*[Batch_Number]
                            data["JobName"]         = data.shape[0]*[ProjectName]
                            data["CrewNumber"]      = data.shape[0]*[Crew_Number]
                            data["Location"]        = data.shape[0]*[ProjectLocation]
                            data["Date"]            = data.shape[0]*[Date]
                            data    = data.loc[:,['BatchNumber','JobName','CrewNumber','Location','Date',
                                                  'Unit_SN','DeviceType','Opened','FaultFound']]
                            data = pd.DataFrame(data)
                            data = data.reset_index(drop=True)
                            dfList.append(data)

                        concatDf = pd.concat(dfList, axis=0, ignore_index =True)
                        data = pd.DataFrame(concatDf)
                        
                        if (data['BatchNumber'].isnull().values.any() == True)|(data['JobName'].isnull().values.any() == True)|(data['CrewNumber'].isnull().values.any() == True)|(data['Location'].isnull().values.any() == True)|(data['Date'].isnull().values.any() == True)|(data['Unit_SN'].isnull().values.any() == True)|(data['DeviceType'].isnull().values.any() == True)|(data['Opened'].isnull().values.any() == True)|(data['FaultFound'].isnull().values.any() == True):
                            tkinter.messagebox.showinfo("Import File Message", "BatchNumber / JobName / CrewNumber / Location / Date / Unit_SN / DeviceType / Opened / FaultFound : Any Value in Any Column Cannot be Empty ")
                        else:
                            tree.delete(*tree.get_children())
                            ImportTotalLBEntries.delete(0,END)
                            ImportDuplicatedEntries.delete(0,END)
                            ImportValidEntries.delete(0,END)
                            conn = sqlite3.connect("Eagle_BadGSRInventoryDatabase.db")
                            cursor = conn.cursor()                        
                            data['Date'] = pd.to_datetime(data['Date']).dt.strftime("%Y-%m-%d")
                            data ['DuplicatedEntries']=data.sort_values(by =['BatchNumber']).duplicated(['BatchNumber','Unit_SN','DeviceType'],keep='last')
                            data_Duplicated = data.loc[data.DuplicatedEntries == True, 'BatchNumber': 'DuplicatedEntries']
                            data_Valid = data.loc[data.DuplicatedEntries == False, 'BatchNumber': 'DuplicatedEntries']
                            data_Duplicated_Count = len(data_Duplicated)
                            data_Valid_Count = len(data_Valid)
                            data.to_sql('Eagle_BadGSRInventoryDatabase_TEMP_IMPORT',conn, if_exists="replace", index=False)
                            for each_rec in range(len(data)):
                                tree.insert("", tk.END, values=list(data.loc[each_rec]))
                            ListBoxTotalImportEntries()
                            ImportDuplicatedEntries.insert(tk.END,data_Duplicated_Count)
                            ImportValidEntries.insert(tk.END,data_Valid_Count)


            def ClearView():
                tree.delete(*tree.get_children())
                ImportTotalLBEntries.delete(0,END)
                ImportDuplicatedEntries.delete(0,END)
                ImportValidEntries.delete(0,END)

            def ListBoxTotalImportEntries():
                ImportTotalLBEntries.delete(0,END)
                Total_count = len(tree.get_children())
                ImportTotalLBEntries.insert(tk.END,Total_count)

            def Exit():
                window.destroy()
                
            def ViewScannedImport():
                tree.delete(*tree.get_children())
                ImportTotalLBEntries.delete(0,END)
                conn = sqlite3.connect("Eagle_BadGSRInventoryDatabase.db")
                Complete_df = pd.read_sql_query("SELECT * FROM Eagle_BadGSRInventoryDatabase_TEMP_IMPORT ORDER BY `Unit_SN` ASC ;", conn)
                data = pd.DataFrame(Complete_df)
                for each_rec in range(len(data)):
                    tree.insert("", tk.END, values=list(data.loc[each_rec]))
                TotalEntries = len(data)
                ImportTotalLBEntries.insert(tk.END,TotalEntries)
                conn.commit()
                conn.close()

            def ViewScannedDuplicatedImport():
                tree.delete(*tree.get_children())
                ImportDuplicatedEntries.delete(0,END)
                conn = sqlite3.connect("Eagle_BadGSRInventoryDatabase.db")
                Complete_df = pd.read_sql_query("SELECT * FROM Eagle_BadGSRInventoryDatabase_TEMP_IMPORT ORDER BY `Unit_SN` ASC ;", conn)
                data = pd.DataFrame(Complete_df)
                data = data.loc[data.DuplicatedEntries == True, 'BatchNumber': 'DuplicatedEntries']
                data = data.reset_index(drop=True)
                for each_rec in range(len(data)):
                    tree.insert("", tk.END, values=list(data.loc[each_rec]))
                TotalDuplicatedEntries = len(data)
                ImportDuplicatedEntries.insert(tk.END,TotalDuplicatedEntries)
                conn.commit()
                conn.close()

            def ViewValidScannedImport():
                tree.delete(*tree.get_children())
                ImportValidEntries.delete(0,END)
                conn = sqlite3.connect("Eagle_BadGSRInventoryDatabase.db")
                Complete_df = pd.read_sql_query("SELECT * FROM Eagle_BadGSRInventoryDatabase_TEMP_IMPORT ORDER BY `Unit_SN` ASC ;", conn)
                data = pd.DataFrame(Complete_df)
                data = data.loc[data.DuplicatedEntries == False, 'BatchNumber': 'DuplicatedEntries']
                data = data.reset_index(drop=True)
                for each_rec in range(len(data)):
                    tree.insert("", tk.END, values=list(data.loc[each_rec]))
                TotalValidEntries = len(data)
                ImportValidEntries.insert(tk.END,TotalValidEntries)
                conn.commit()
                conn.close()
                                            
            def Submit_Scanned_import_ToMasterDB():
                iSubmit = tkinter.messagebox.askyesno("Valid Entries Submit to Master DB", "Confirm if you want to submit only Valid Entries to MasterDB")
                if iSubmit >0:
                    conn = sqlite3.connect("Eagle_BadGSRInventoryDatabase.db")
                    Complete_df = pd.read_sql_query("SELECT * FROM Eagle_BadGSRInventoryDatabase_TEMP_IMPORT ORDER BY `Unit_SN` ASC ;", conn)
                    cur=conn.cursor()
                    data = pd.DataFrame(Complete_df)
                    data = data.loc[data.DuplicatedEntries == False, 'BatchNumber': 'FaultFound']
                    data = data.reset_index(drop=True)
                    data.to_sql('Eagle_BadGSRInventoryDatabase',conn, if_exists="append", index=False)    
                    tkinter.messagebox.showinfo("Submitted to Inventory Database(DB)","You have Submitted a Record to Inventory Database(DB)")
                    ClearView()
                    cur.execute("DELETE FROM Eagle_BadGSRInventoryDatabase_TEMP_IMPORT")
                    conn.commit()
                    conn.close()
                    UpdateImportedScannedToMASTERDB()
                    return

            
            def UpdateImportedScannedToMASTERDB():
                conn = sqlite3.connect("Eagle_BadGSRInventoryDatabase.db")
                Complete_df = pd.read_sql_query("SELECT * FROM Eagle_BadGSRInventoryDatabase ORDER BY `Unit_SN` ASC ;", conn)
                data = pd.DataFrame(Complete_df)
                data ['DuplicatedEntries']=data.sort_values(by =['BatchNumber']).duplicated(['BatchNumber','Unit_SN','DeviceType'],keep='last')

                data_Duplicated = data.loc[data.DuplicatedEntries == True, 'BatchNumber': 'FaultFound']
                data_View_Duplicated = data_Duplicated.reset_index(drop=True)
                data_View_Duplicated.to_sql('Eagle_BadGSRInventoryDatabase_TEMP_DUPLICATED',conn, if_exists="replace", index=False)
                
                data = data.loc[data.DuplicatedEntries == False, 'BatchNumber': 'FaultFound']
                data = data.reset_index(drop=True)
                self.cur=conn.cursor()
                data.to_sql('Eagle_BadGSRInventoryDatabase',conn, if_exists="replace", index=False)                    
                conn.commit()
                conn.close()
            

            ## Labels 
            L1Import_Total = Label(window, text = "A: Total Import:-", font=("arial", 10,'bold'),bg = "ghost white").place(x=874,y=10)
            L1Import_Duplicated = Label(window, text = "B: Duplicated:-", font=("arial", 10,'bold'),bg = "ghost white").place(x=874,y=100)
            L1Import_Valid = Label(window, text = "C: Valid Entry:-", font=("arial", 10,'bold'),bg = "ghost white").place(x=874,y=190)
            
            ## Entries Field
            ImportTotalLBEntries  = Entry(window, font=('aerial', 12, 'bold'),textvariable= IntVar(), width = 8)
            ImportTotalLBEntries.place(x=885,y=33)

            ImportDuplicatedEntries  = Entry(window, font=('aerial', 12, 'bold'),textvariable= IntVar(), width = 8)
            ImportDuplicatedEntries.place(x=885,y=125)

            ImportValidEntries  = Entry(window, font=('aerial', 12, 'bold'),textvariable= IntVar(), width = 8)
            ImportValidEntries.place(x=885,y=217)

            ## Buttons
            btnImportView = Button(window, text="View All Import", font=('aerial', 9, 'bold'), height =1, width=13, bd=2, bg= 'grey', command = ViewScannedImport)
            btnImportView.place(x=874,y=63)

            btnDuplicatedView = Button(window, text="View Duplicated", font=('aerial', 9, 'bold'), height =1, width=13, bd=2, bg= 'grey', command = ViewScannedDuplicatedImport)
            btnDuplicatedView.place(x=874,y=154)

            btnValidView = Button(window, text="View Valid Entry", font=('aerial', 9, 'bold'), height =1, width=13, bd=2, bg= 'grey', command = ViewValidScannedImport)
            btnValidView.place(x=874,y=245)
            
            btnImport = Button(window, text="Import Scaned File", font=('aerial', 9, 'bold'), height =1, width=16, bd=2, bg= 'grey', command = ImportBatchScannedFile)
            btnImport.place(x=874,y=390)

            btnSubmit = Button(window, text="Submit Valid to DB", font=('aerial', 9, 'bold'), height =1, width=16, bd=2,bg= 'grey', command = Submit_Scanned_import_ToMasterDB)
            btnSubmit.place(x=874,y=425)

            btnClearView = Button(window, text="Clear View", font=('aerial', 9, 'bold'), height =1, width=12, bd=2,bg= 'grey', command =ClearView)
            btnClearView.place(x=874,y=466)

            btnExit = Button(window, text="Exit", font=('aerial', 9, 'bold'), height =1, width=12, bd=2,bg= 'grey', command = Exit)
            btnExit.place(x=874,y=496)

        def ImportMasterDBFile():
            window = Tk()
            window.title("Import Master DB File Viewer")
            window.config(bg="ghost white")
            width = 995
            height = 540
            screen_width = window.winfo_screenwidth()
            screen_height = window.winfo_screenheight()
            x = (screen_width/2) - (width/2)
            y = (screen_height/2) - (height/2)
            window.geometry("%dx%d+%d+%d" % (width, height, x, y))
            window.resizable(0, 0)
            TableMargin = Frame(window)
            TableMargin.pack(side=TOP)
            TableMargin.pack(side=LEFT)
            scrollbarx = Scrollbar(TableMargin, orient=HORIZONTAL)
            scrollbary = Scrollbar(TableMargin, orient=VERTICAL)
            tree = ttk.Treeview(TableMargin, column=("column1", "column2", "column3", "column4", "column5", "column6", "column7", "column8", "column9", "column10"),
                                        height=25, show='headings')
            scrollbary.config(command=tree.yview)
            scrollbary.pack(side=RIGHT, fill=Y)
            scrollbarx.config(command=tree.xview)
            scrollbarx.pack(side=BOTTOM, fill=X)
            tree.heading("#1", text="BatchNumber", anchor=W)
            tree.heading("#2", text="JobName", anchor=W)
            tree.heading("#3", text="CrewNumber", anchor=W)
            tree.heading("#4", text="Location", anchor=W)
            tree.heading("#5", text="Date", anchor=W)            
            tree.heading("#6", text="Unit_SN", anchor=W)
            tree.heading("#7", text="DeviceType" ,anchor=W)
            tree.heading("#8", text="Opened", anchor=W)
            tree.heading("#9", text="FaultFound", anchor=W)
            tree.heading("#10", text="Duplicated", anchor=W)
            tree.column('#1', stretch=NO, minwidth=0, width=90)            
            tree.column('#2', stretch=NO, minwidth=0, width=100)
            tree.column('#3', stretch=NO, minwidth=0, width=90)
            tree.column('#4', stretch=NO, minwidth=0, width=80)
            tree.column('#5', stretch=NO, minwidth=0, width=70)
            tree.column('#6', stretch=NO, minwidth=0, width=80)
            tree.column('#7', stretch=NO, minwidth=0, width=80)
            tree.column('#8', stretch=NO, minwidth=0, width=80)
            tree.column('#9', stretch=NO, minwidth=0, width=100)
            tree.column('#10', stretch=NO, minwidth=0, width=80)
            tree.pack()

            self.filename = None
            self.df = None

            def ImportDBMaster():
                name = askopenfilename(filetypes=[('CSV File', '*.csv'), ('Excel File', ('*.xls', '*.xlsx'))])
                if name:
                    if name.endswith('.csv'):
                        self.df = pd.read_csv(name, header = None, skiprows = {0})
                        self.df.rename(columns = {0:'BatchNumber', 1:'JobName', 2:'CrewNumber', 3:'Location', 4:'Date',
                                                  5: 'Unit_SN', 6:'DeviceType', 7:'Opened', 8:'FaultFound'},inplace = True)
                    else:
                        self.df = pd.read_excel(name, header = None, skiprows = {0})
                        self.df.rename(columns = {0:'BatchNumber', 1:'JobName', 2:'CrewNumber', 3:'Location', 4:'Date',
                                                  5: 'Unit_SN', 6:'DeviceType', 7:'Opened', 8:'FaultFound'},inplace = True)                        
                    data = pd.DataFrame(self.df)

                    if (data['BatchNumber'].isnull().values.any() == True)|(data['JobName'].isnull().values.any() == True)|(data['CrewNumber'].isnull().values.any() == True)|(data['Location'].isnull().values.any() == True)|(data['Date'].isnull().values.any() == True)|(data['Unit_SN'].isnull().values.any() == True)|(data['DeviceType'].isnull().values.any() == True)|(data['Opened'].isnull().values.any() == True)|(data['FaultFound'].isnull().values.any() == True):
                        tkinter.messagebox.showinfo("Import File Message", "BatchNumber / JobName / CrewNumber / Location / Date / Unit_SN / DeviceType / Opened / FaultFound : Any Value in Any Column Cannot be Empty ")
                    else:
                        tree.delete(*tree.get_children())
                        ImportTotalLBEntries.delete(0,END)
                        ImportDuplicatedEntries.delete(0,END)
                        ImportValidEntries.delete(0,END)
                        conn = sqlite3.connect("Eagle_BadGSRInventoryDatabase.db")
                        cursor = conn.cursor()                        
                        data['Date'] = pd.to_datetime(data['Date']).dt.strftime("%Y-%m-%d")
                        data ['DuplicatedEntries']=data.sort_values(by =['BatchNumber']).duplicated(['BatchNumber','Unit_SN','DeviceType'],keep='last')
                        data_Duplicated = data.loc[data.DuplicatedEntries == True, 'BatchNumber': 'DuplicatedEntries']
                        data_Valid = data.loc[data.DuplicatedEntries == False, 'BatchNumber': 'DuplicatedEntries']
                        data_Duplicated_Count = len(data_Duplicated)
                        data_Valid_Count = len(data_Valid)
                        data.to_sql('Eagle_BadGSRInventoryDatabase_TEMP_IMPORT',conn, if_exists="replace", index=False)
                        for each_rec in range(len(data)):
                            tree.insert("", tk.END, values=list(data.loc[each_rec]))
                        ListBoxTotalImportEntries()
                        ImportDuplicatedEntries.insert(tk.END,data_Duplicated_Count)
                        ImportValidEntries.insert(tk.END,data_Valid_Count)

            def ListBoxTotalImportEntries():
                ImportTotalLBEntries.delete(0,END)
                Total_count = len(tree.get_children())
                ImportTotalLBEntries.insert(tk.END,Total_count)

            def ClearView():
                tree.delete(*tree.get_children())
                ImportTotalLBEntries.delete(0,END)
                ImportDuplicatedEntries.delete(0,END)
                ImportValidEntries.delete(0,END)

            def Exit():
                window.destroy()
                
            def ViewDBMasterImport():
                tree.delete(*tree.get_children())
                ImportTotalLBEntries.delete(0,END)
                conn = sqlite3.connect("Eagle_BadGSRInventoryDatabase.db")
                Complete_df = pd.read_sql_query("SELECT * FROM Eagle_BadGSRInventoryDatabase_TEMP_IMPORT ORDER BY `Unit_SN` ASC ;", conn)
                data = pd.DataFrame(Complete_df)
                for each_rec in range(len(data)):
                    tree.insert("", tk.END, values=list(data.loc[each_rec]))
                TotalEntries = len(data)
                ImportTotalLBEntries.insert(tk.END,TotalEntries)
                conn.commit()
                conn.close()

            def ViewDBMasterDuplicatedImport():
                tree.delete(*tree.get_children())
                ImportDuplicatedEntries.delete(0,END)
                conn = sqlite3.connect("Eagle_BadGSRInventoryDatabase.db")
                Complete_df = pd.read_sql_query("SELECT * FROM Eagle_BadGSRInventoryDatabase_TEMP_IMPORT ORDER BY `Unit_SN` ASC ;", conn)
                data = pd.DataFrame(Complete_df)
                data = data.loc[data.DuplicatedEntries == True, 'BatchNumber': 'DuplicatedEntries']
                data = data.reset_index(drop=True)
                for each_rec in range(len(data)):
                    tree.insert("", tk.END, values=list(data.loc[each_rec]))
                TotalDuplicatedEntries = len(data)
                ImportDuplicatedEntries.insert(tk.END,TotalDuplicatedEntries)
                conn.commit()
                conn.close()

            def ViewValidDBMasterImport():
                tree.delete(*tree.get_children())
                ImportValidEntries.delete(0,END)
                conn = sqlite3.connect("Eagle_BadGSRInventoryDatabase.db")
                Complete_df = pd.read_sql_query("SELECT * FROM Eagle_BadGSRInventoryDatabase_TEMP_IMPORT ORDER BY `Unit_SN` ASC ;", conn)
                data = pd.DataFrame(Complete_df)
                data = data.loc[data.DuplicatedEntries == False, 'BatchNumber': 'DuplicatedEntries']
                data = data.reset_index(drop=True)
                for each_rec in range(len(data)):
                    tree.insert("", tk.END, values=list(data.loc[each_rec]))
                TotalValidEntries = len(data)
                ImportValidEntries.insert(tk.END,TotalValidEntries)
                conn.commit()
                conn.close()
                                            
            def Submit_import_data_MasterDB():
                iSubmit = tkinter.messagebox.askyesno("Valid Entries Submit to Master DB", "Confirm if you want to submit only Valid Entries to MasterDB")
                if iSubmit >0:
                    conn = sqlite3.connect("Eagle_BadGSRInventoryDatabase.db")
                    Complete_df = pd.read_sql_query("SELECT * FROM Eagle_BadGSRInventoryDatabase_TEMP_IMPORT ORDER BY `Unit_SN` ASC ;", conn)
                    cur=conn.cursor()
                    data = pd.DataFrame(Complete_df)
                    data = data.loc[data.DuplicatedEntries == False, 'BatchNumber': 'FaultFound']
                    data = data.reset_index(drop=True)
                    data.to_sql('Eagle_BadGSRInventoryDatabase',conn, if_exists="append", index=False)    
                    tkinter.messagebox.showinfo("Submitted to Inventory Database(DB)","You have Submitted a Record to Inventory Database(DB)")
                    ClearView()
                    cur.execute("DELETE FROM Eagle_BadGSRInventoryDatabase_TEMP_IMPORT")
                    conn.commit()
                    conn.close()
                    UpdateImportedMasterDBToMASTERDB()
                    window.destroy()
                    return
            
            def UpdateImportedMasterDBToMASTERDB():
                conn = sqlite3.connect("Eagle_BadGSRInventoryDatabase.db")
                Complete_df = pd.read_sql_query("SELECT * FROM Eagle_BadGSRInventoryDatabase ORDER BY `Unit_SN` ASC ;", conn)
                data = pd.DataFrame(Complete_df)
                data ['DuplicatedEntries']=data.sort_values(by =['BatchNumber']).duplicated(['BatchNumber','Unit_SN','DeviceType'],keep='last')

                data_Duplicated = data.loc[data.DuplicatedEntries == True, 'BatchNumber': 'FaultFound']
                data_View_Duplicated = data_Duplicated.reset_index(drop=True)
                data_View_Duplicated.to_sql('Eagle_BadGSRInventoryDatabase_TEMP_DUPLICATED',conn, if_exists="replace", index=False)
                
                data = data.loc[data.DuplicatedEntries == False, 'BatchNumber': 'FaultFound']
                data = data.reset_index(drop=True)
                self.cur=conn.cursor()
                data.to_sql('Eagle_BadGSRInventoryDatabase',conn, if_exists="replace", index=False)                    
                conn.commit()
                conn.close()
            
            ## Labels 
            L1Import_Total = Label(window, text = "A: Total Import:-", font=("arial", 10,'bold'),bg = "ghost white").place(x=874,y=10)
            L1Import_Duplicated = Label(window, text = "B: Duplicated:-", font=("arial", 10,'bold'),bg = "ghost white").place(x=874,y=100)
            L1Import_Valid = Label(window, text = "C: Valid Entry:-", font=("arial", 10,'bold'),bg = "ghost white").place(x=874,y=190)
            
            ## Entries Field
            ImportTotalLBEntries  = Entry(window, font=('aerial', 12, 'bold'),textvariable= IntVar(), width = 8)
            ImportTotalLBEntries.place(x=885,y=33)

            ImportDuplicatedEntries  = Entry(window, font=('aerial', 12, 'bold'),textvariable= IntVar(), width = 8)
            ImportDuplicatedEntries.place(x=885,y=125)

            ImportValidEntries  = Entry(window, font=('aerial', 12, 'bold'),textvariable= IntVar(), width = 8)
            ImportValidEntries.place(x=885,y=217)

            ## Buttons
            btnImportView = Button(window, text="View All Import", font=('aerial', 9, 'bold'), height =1, width=13, bd=2, bg= 'grey', command =ViewDBMasterImport)
            btnImportView.place(x=874,y=63)

            btnDuplicatedView = Button(window, text="View Duplicated", font=('aerial', 9, 'bold'), height =1, width=13, bd=2, bg= 'grey',command =ViewDBMasterDuplicatedImport)
            btnDuplicatedView.place(x=874,y=154)

            btnValidView = Button(window, text="View Valid Entry", font=('aerial', 9, 'bold'), height =1, width=13, bd=2, bg= 'grey',command = ViewValidDBMasterImport)
            btnValidView.place(x=874,y=245)
            
            btnImport = Button(window, text="Import Inventory File", font=('aerial', 9, 'bold'), height =1, width=16, bd=2, bg= 'grey',command = ImportDBMaster)
            btnImport.place(x=874,y=390)

            btnSubmit = Button(window, text="Submit Valid to DB", font=('aerial', 9, 'bold'), height =1, width=16, bd=2,bg= 'grey', command = Submit_import_data_MasterDB)
            btnSubmit.place(x=874,y=425)

            btnClearView = Button(window, text="Clear View", font=('aerial', 9, 'bold'), height =1, width=12, bd=2,bg= 'grey', command = ClearView)
            btnClearView.place(x=874,y=466)

            btnExit = Button(window, text="Exit", font=('aerial', 9, 'bold'), height =1, width=12, bd=2,bg= 'grey', command = Exit)
            btnExit.place(x=874,y=496)
           

        #----------------- Frames-------------
        menu = Menu(self.root)
        self.root.config(menu=menu)
        filemenu = Menu(menu, tearoff=0)
        menu.add_cascade(label="File", menu=filemenu)
        filemenu.add_command(label="Import Master DB File", command = ImportMasterDBFile)
        filemenu.add_command(label="Export Master DB File", command=ExportCompleteDB)
        filemenu.add_separator()
        filemenu.add_command(label="Exit", command=iExit)

        filemenu1 = Menu(menu, tearoff=0)
        menu.add_cascade(label="Transmittal Out", menu=filemenu1)
        filemenu1.add_command(label="Generate Transmittal Out", command = GenerateTransmittalOut)
        filemenu1.add_command(label="Transmittal Receiver Location", command = GenerateReceivingAddress.Generate_ReceiverAddress)
        filemenu1.add_command(label="PDF Convert Transmittal Out", command = Transmittal_ConvertToPDF)
        
        TitFrame = Frame(self.root, bd = 2, padx= 5, pady= 4, bg = "#006dcc", relief = RIDGE)
        TitFrame.pack(side = TOP)

        L1 = Label(self.root, text = "A: Eagle Bad Unit Entry", font=("arial", 10,'bold'),bg = "green").place(x=6,y=14)
        L2 = Label(self.root, text = "B: Eagle Bad Unit Details", font=("arial", 10,'bold'),bg = "green").place(x=480,y=14)

        
        DataFrameLEFT = LabelFrame(self.root, bd = 2, width = 520, height = 700, padx= 6, pady= 10,relief = RIDGE,
                                   bg = "Ghost White",font=('aerial', 15, 'bold'))
        DataFrameLEFT.place(x=4,y=40)

                #----------------- Tree View Frames-------------
        TableMargin = Frame(self.root, relief = RIDGE)
        TableMargin.pack(side=BOTTOM)
        TableMargin.pack(side=RIGHT)
        scrollbarx = Scrollbar(TableMargin, orient=HORIZONTAL)
        scrollbary = Scrollbar(TableMargin, orient=VERTICAL)
        tree = ttk.Treeview(TableMargin, column=("column1", "column2", "column3", "column4", "column5", "column6", "column7", "column8",  "column9"),
                            height=28, show='headings')
        scrollbary.config(command=tree.yview)
        scrollbary.pack(side=RIGHT, fill=Y)
        scrollbarx.config(command=tree.xview)
        scrollbarx.pack(side=BOTTOM, fill=X)        
        tree.heading("#1", text="BatchNumber", anchor=W)
        tree.heading("#2", text="JobName", anchor=W)
        tree.heading("#3", text="Crew", anchor=W)
        tree.heading("#4", text="Location", anchor=W)
        tree.heading("#5", text="Date", anchor=W)            
        tree.heading("#6", text="Unit_SN", anchor=W)
        tree.heading("#7", text="DeviceType" ,anchor=W)        
        tree.heading("#8", text="Opened", anchor=W)
        tree.heading("#9", text="FaultFound", anchor=W)
        
        tree.column('#1', stretch=NO, minwidth=0, width=120)            
        tree.column('#2', stretch=NO, minwidth=0, width=180)
        tree.column('#3', stretch=NO, minwidth=0, width=50)
        tree.column('#4', stretch=NO, minwidth=0, width=110)
        tree.column('#5', stretch=NO, minwidth=0, width=80)
        tree.column('#6', stretch=NO, minwidth=0, width=80)
        tree.column('#7', stretch=NO, minwidth=0, width=80)
        tree.column('#8', stretch=NO, minwidth=0, width=60)
        tree.column('#9', stretch=NO, minwidth=0, width=100)
        
        style = ttk.Style()
        style.theme_use("clam")
        style.configure(".", font=('aerial', 9), foreground="black")
        style.configure("Treeview", foreground='black')
        style.configure("Treeview.Heading",font=('aerial', 9,'bold'), background='Ghost White', foreground='blue',fieldbackground='Ghost White')

        Treepopup = Menu(tree, tearoff=0)
        Treepopup.add_command(label="Delete Selected Entries", command=DeleteData)        
        Treepopup.add_command(label="Update Selected Batch Number", command=UpdateSlectedBatchNumber)
        Treepopup.add_command(label="Update Selected Fault Found", command=UpdateSlectedFaultFound)
        Treepopup.add_command(label="Update Selected Job Name", command=UpdateSlectedJobName)
        Treepopup.add_command(label="Export Selected ListBox Entries", command=ExportSelectedListBoxView)
        Treepopup.add_separator()
        Treepopup.add_command(label="Exit", command=iExit)

        def Treepopup_do_popup(event):
            try:
                Treepopup.tk_popup(event.x_root, event.y_root, 0)
            finally:
                Treepopup.grab_release()

        tree.bind("<Button-3>", Treepopup_do_popup)
        tree.pack()
    

##        #----------------- Labels and Entry Wizard------------

        self.lblBatchNumber = Label(DataFrameLEFT, font=('aerial', 10, 'bold'), text = "1. Batch Number:", padx =10, pady= 10, bg = "Ghost White")
        self.lblBatchNumber.grid(row =0, column = 0, sticky =W)
        self.txtBatchNumber = ttk.Combobox(DataFrameLEFT, font=('aerial', 10, 'bold'), textvariable = BatchNumber, width = 30)
        self.txtBatchNumber.grid(row =0, column = 1)
        self.txtBatchNumber['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_BatchNumber())))

        self.lblJobName = Label(DataFrameLEFT, font=('aerial', 10, 'bold'), text = "2. Job Name:", padx =10, pady= 10, bg = "Ghost White")
        self.lblJobName.grid(row =1, column = 0, sticky =W)
        self.txtJobName = ttk.Combobox(DataFrameLEFT, font=('aerial', 10, 'bold'), textvariable = JobName, width = 30)
        self.txtJobName.grid(row =1, column = 1)
        self.txtJobName['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_JobName())))

        self.lblCrewNumber = Label(DataFrameLEFT, font=('aerial', 10, 'bold'), text = "3. Crew Number:", padx =10, pady= 10, bg = "Ghost White")
        self.lblCrewNumber.grid(row =2, column = 0, sticky =W)
        self.txtCrewNumber = ttk.Combobox(DataFrameLEFT, font=('aerial', 10, 'bold'), textvariable = CrewNumber, width = 30)
        self.txtCrewNumber.grid(row =2, column = 1)
        self.txtCrewNumber['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_CrewNumber())))
        
        self.lblLocation = Label(DataFrameLEFT, font=('aerial', 10, 'bold'), text = "4. Crew Location:", padx =10, pady= 10, bg = "Ghost White")
        self.lblLocation.grid(row =3, column = 0, sticky =W)
        self.txtLocation = ttk.Combobox(DataFrameLEFT, font=('aerial', 10, 'bold'), textvariable = Location, width = 30)
        self.txtLocation.grid(row =3, column = 1)
        self.txtLocation['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_Location())))
        
        self.lblDate = Label(DataFrameLEFT, font=('aerial', 10, 'bold'), text = "5. Date Reported (yyyy-mm-dd):", padx =10, pady= 10, bg = "Ghost White")
        self.lblDate.grid(row =4, column = 0, sticky =W)
        self.txtDate = ttk.Combobox(DataFrameLEFT, font=('aerial', 10, 'bold'), textvariable = Date, width = 30)
        self.txtDate.grid(row =4, column = 1)
        self.txtDate['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_Date())))

        self.lblUnit_SN = Label(DataFrameLEFT, font=('aerial', 10, 'bold'), text = "6. Unit Serial Number:", padx =10, pady= 10, bg = "Ghost White")
        self.lblUnit_SN.grid(row =5, column = 0, sticky =W)
        self.txtUnit_SN = ttk.Combobox(DataFrameLEFT, font=('aerial', 10, 'bold'), textvariable = Unit_SN, width = 30)
        self.txtUnit_SN.grid(row =5, column = 1)
        self.txtUnit_SN['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_Unit_SN())))

        self.lblDeviceType = Label(DataFrameLEFT, font=('aerial', 10, 'bold'), text = "7. Device Type:", padx =10 , pady= 10, bg = "Ghost White")
        self.lblDeviceType.grid(row =6, column = 0, sticky =W)
        self.txtDeviceType = ttk.Combobox(DataFrameLEFT, font=('aerial', 10, 'bold'), textvariable = DeviceType, width = 30)
        self.txtDeviceType.grid(row =6, column = 1)
        self.txtDeviceType['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_DeviceType())))

        self.lblOpened = Label(DataFrameLEFT, font=('aerial', 10, 'bold'), text = "9. Box Opened (Yes/No):", padx =10, pady= 10, bg = "Ghost White")
        self.lblOpened.grid(row =7, column = 0, sticky =W)
        self.txtOpened = ttk.Combobox(DataFrameLEFT, font=('aerial', 10, 'bold'), textvariable = Opened, width = 30)
        self.txtOpened.grid(row =7, column = 1)
        self.txtOpened['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_Opened())))

        self.lblFaultFound = Label(DataFrameLEFT, font=('aerial', 10, 'bold'), text = "8. Fault Found:", padx =10, pady= 10, bg = "Ghost White")
        self.lblFaultFound.grid(row =8, column = 0, sticky =W)
        self.txtFaultFound = ttk.Combobox(DataFrameLEFT, font=('aerial', 10, 'bold'), textvariable = FaultFound, width = 30, height =2)
        self.txtFaultFound.grid(row =8, column = 1)
        self.txtFaultFound['values'] = sorted(list(set(Eagle_BadGSRInventoryDatabase_BackEnd.Combo_input_FaultFound())))

        
##        #----------------- Tree View Select Event------------
        
        tree.bind('<<TreeviewSelect>>',InventoryRec)



##        #----------------- Button Widget------------
       
        btnUpdateMasterDB = Button(self.root, text="Update Master DB", font=('aerial', 10, 'bold'), height =1, width=15, bd=2,command = UpdateMasterDB )
        btnUpdateMasterDB.place(x=5,y=425)

        btnModifyUpdateData = Button(self.root, text="Modify Entry", font=('aerial', 10, 'bold'), height =1, width=11, bd=2, command = update)
        btnModifyUpdateData.place(x=155,y=425)

        btnSearchData = Button(self.root, text="Search Entry", font=('aerial', 10, 'bold'), height =1, width=10, bd=2, command = searchDatabase)
        btnSearchData.place(x=260,y=425)

        btnAddData = Button(self.root, text="Add New Entry", font=('aerial', 10, 'bold'), height =1, width=13, bd=2,command = AddData )
        btnAddData.place(x=357,y=425)

        btnClearData = Button(self.root, text="Clear Entry", font=('aerial', 9, 'bold'), height =1, width=10, bd=2, command = ClearData)
        btnClearData.place(x=392,y=11)


        btnViewMasterDB = Button(self.root, text="Display Master DB", font=('aerial', 10, 'bold'), bg='orange', height =1, width=15, bd=4, command = ViewMasterDB)
        btnViewMasterDB.place(x=1125,y=7)
        self.txtTotalEntries  = Entry(self.root, font=('aerial', 12, 'bold'),textvariable=TOTALE, width = 9, bd=4)
        self.txtTotalEntries.place(x=1262,y=8)

        btnKeySearchListbox = Button(self.root, text="Search by Keyword", font=('aerial', 9, 'bold'), height =1, width=16, bd=4, command = KeySearch)
        btnKeySearchListbox.place(x=710,y=9)
        self.txtKeySearch  = Entry(self.root, font=('aerial', 12, 'bold'),bd=4, textvariable=SEARCH, width = 14)
        self.txtKeySearch.place(x=840,y=9)

        btnViewDuplicated = Button(self.root, text="View Duplicated", font=('aerial', 10, 'bold'), height =1, width=13, bd=2,command = ViewDuplicated)
        btnViewDuplicated.place(x=477,y=650)
        self.txtDuplicatedEntries  = Entry(self.root, font=('aerial', 12, 'bold'),textvariable=TOTALD, width = 4, bd=2)
        self.txtDuplicatedEntries.place(x=595,y=650)

        btnImportFile = Button(self.root, text="Import Crew \n Batch Scanned File", font=('aerial', 10, 'bold'), height =2, width=16, bd=2, command = ImportBatchScannedFile)
        btnImportFile.place(x=800,y=650)

        btnDeleteListBox = Button(self.root, text="Delete Selected", font=('aerial', 10, 'bold'), height =1, width=13, bd=2, command = DeleteData)
        btnDeleteListBox.place(x=1005,y=650)

        btnExportListBox = Button(self.root, text="Export ListBox", font=('aerial', 10, 'bold'), height =1, width=12, bd=2, command = ExportListboxEntries)
        btnExportListBox.place(x=1123,y=650)

        btnClearClearAllView = Button(self.root, text="ClearView", font=('aerial', 10, 'bold'), height =1, width=9, bd=2, command = ClearAllView)
        btnClearClearAllView.place(x=1231,y=650)

        btnExitData = Button(self.root, text="Exit", font=('aerial', 10, 'bold'), height =1, width=4, bd=2, command = iExit)
        btnExitData.place(x=1314,y=650)




if __name__ == '__main__':
    root = Tk()
    application  = Inventory (root)
    root.mainloop()

